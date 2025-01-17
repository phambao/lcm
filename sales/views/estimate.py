from datetime import datetime
from datetime import timedelta
import json
import re

from django.utils.timezone import now
from django.apps import apps
from django.db.models import Value, Q, Subquery
from django_filters.filters import _truncate
from openpyxl import Workbook, load_workbook
from rest_framework import generics, permissions, status, filters as rf_filters
from django_filters import rest_framework as filters
from rest_framework.decorators import api_view, permission_classes
from rest_framework.generics import get_object_or_404
from rest_framework.pagination import LimitOffsetPagination
from rest_framework.response import Response
from rest_framework import serializers

from base.permissions import EstimatePermissions
from base.utils import file_response
from sales.filters.estimate import FormulaFilter, EstimateTemplateFilter, AssembleFilter, GroupFormulaFilter,\
    DescriptionFilter, UnitFilter, DataEntryFilter
from sales.models import DataPoint, Catalog
from sales.models.estimate import POFormula, POFormulaGrouping, DataEntry, POFormulaToDataEntry, UnitLibrary, \
    DescriptionLibrary, Assemble, EstimateTemplate, MaterialView
from sales.serializers.catalog import CatalogEstimateSerializer, CatalogEstimateWithParentSerializer, CatalogLevelValueSerializer, CatalogSerializer
from sales.serializers.estimate import POFormulaGroupCompactSerializer, POFormulaSerializer, POFormulaGroupingSerializer, DataEntrySerializer, \
    UnitLibrarySerializer, DescriptionLibrarySerializer, LinkedDescriptionSerializer, AssembleSerializer, \
    EstimateTemplateSerializer, TaggingSerializer, GroupFormulasSerializer, POFormulaCompactSerializer, \
    AssembleCompactSerializer, EstimateTemplateCompactSerializer
from sales.views.catalog import parse_c_table
from api.middleware import get_request
from base.views.base import CompanyFilterMixin
from base.constants import null, true, false

DATA_ENTRY_PREFETCH_RELATED = ['data_entry__unit', 'data_entry__material_selections', 'data_entry']
FORMULA_PREFETCH_RELATED = ['self_data_entries__' + i for i in DATA_ENTRY_PREFETCH_RELATED]
GROUP_PREFETCH_RELATED = ['group_formulas__' + i for i in FORMULA_PREFETCH_RELATED]
ASSEMBLE_PREFETCH_RELATED = ['assemble_formulas__' + i for i in FORMULA_PREFETCH_RELATED]
ESTIMATE_PREFETCH_RELATED = ['assembles__' + i for i in ASSEMBLE_PREFETCH_RELATED]
ESTIMATE_DATA_ENTRY_PREFETCH_RELATED = ['data_entries__' + i for i in DATA_ENTRY_PREFETCH_RELATED]
MATERIAL_PREFETCH_RELATED = ['material_views__' + i for i in DATA_ENTRY_PREFETCH_RELATED]
ALL_ESTIMATE_PREFETCH_RELATED = [*ESTIMATE_PREFETCH_RELATED, *MATERIAL_PREFETCH_RELATED,
                                 *ESTIMATE_DATA_ENTRY_PREFETCH_RELATED, 'data_views', 'data_views__unit']
ProposalWriting = apps.get_model('sales', 'ProposalWriting')


class POFormulaList(CompanyFilterMixin, generics.ListCreateAPIView):
    queryset = POFormula.objects.filter(is_show=True, group=None).\
        prefetch_related(*FORMULA_PREFETCH_RELATED).select_related('assemble', 'group').order_by('-modified_date')
    serializer_class = POFormulaSerializer
    permission_classes = [permissions.IsAuthenticated & EstimatePermissions]
    filter_backends = (filters.DjangoFilterBackend, rf_filters.SearchFilter)
    filterset_class = FormulaFilter
    search_fields = ('name', )


class POFormulaReMarkOnGroupList(POFormulaList):
    queryset = POFormula.objects.filter(is_show=True).\
        prefetch_related(*FORMULA_PREFETCH_RELATED).select_related('assemble', 'group').order_by('-modified_date')
    search_fields = ('name', 'formula', 'quantity', 'markup', 'charge', 'material', 'cost',
                     'unit', 'gross_profit', 'description_of_formula')


class POFormulaCompactList(POFormulaList):
    serializer_class = POFormulaCompactSerializer


class POFormulaDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = POFormula.objects.all().prefetch_related(*FORMULA_PREFETCH_RELATED)
    serializer_class = POFormulaSerializer
    permission_classes = [permissions.IsAuthenticated & EstimatePermissions]


class POFormulaGroupingList(CompanyFilterMixin, generics.ListCreateAPIView):
    queryset = POFormulaGrouping.objects.all().select_related().order_by('-modified_date').distinct()
    serializer_class = POFormulaGroupingSerializer
    permission_classes = [permissions.IsAuthenticated & EstimatePermissions]
    filter_backends = (filters.DjangoFilterBackend,)
    filterset_class = GroupFormulaFilter


class POFormulaGroupCompactList(POFormulaGroupingList):
    serializer_class = POFormulaGroupCompactSerializer


class POFormulaGroupingCreate(CompanyFilterMixin, generics.CreateAPIView):
    queryset = POFormulaGrouping.objects.all().order_by('-modified_date').distinct()
    serializer_class = GroupFormulasSerializer
    permission_classes = [permissions.IsAuthenticated & EstimatePermissions]


class POFormulaGroupingDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = POFormulaGrouping.objects.all().prefetch_related(*GROUP_PREFETCH_RELATED)
    serializer_class = POFormulaGroupingSerializer
    permission_classes = [permissions.IsAuthenticated & EstimatePermissions]


class DataEntryList(CompanyFilterMixin, generics.ListCreateAPIView):
    queryset = DataEntry.objects.filter(is_show=True).prefetch_related('material_selections', 'unit').select_related('unit').order_by('-modified_date')
    serializer_class = DataEntrySerializer
    permission_classes = [permissions.IsAuthenticated & EstimatePermissions]
    filter_backends = (filters.DjangoFilterBackend, rf_filters.SearchFilter)
    filterset_class = DataEntryFilter
    search_fields = ('name', )


class DataEntryDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = DataEntry.objects.all().prefetch_related('material_selections', 'unit').select_related('unit')
    serializer_class = DataEntrySerializer
    permission_classes = [permissions.IsAuthenticated & EstimatePermissions]

    def destroy(self, request, *args, **kwargs):
        pk = self.kwargs.get('pk')
        if POFormulaToDataEntry.objects.filter(data_entry__pk=pk).exists():
            raise serializers.ValidationError({'error': 'This data entry have been used in other formulas'})
        return super().destroy(request, *args, **kwargs)


class UnitLibraryList(CompanyFilterMixin, generics.ListCreateAPIView):
    queryset = UnitLibrary.objects.all().order_by('-modified_date')
    serializer_class = UnitLibrarySerializer
    permission_classes = [permissions.IsAuthenticated & EstimatePermissions]
    filter_backends = (filters.DjangoFilterBackend, rf_filters.SearchFilter)
    filterset_class = UnitFilter
    search_fields = ('name', )


class UnitLibraryDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = UnitLibrary.objects.all()
    serializer_class = UnitLibrarySerializer
    permission_classes = [permissions.IsAuthenticated & EstimatePermissions]


class DescriptionLibraryList(CompanyFilterMixin, generics.ListCreateAPIView):
    queryset = DescriptionLibrary.objects.all().order_by('-modified_date')
    serializer_class = DescriptionLibrarySerializer
    permission_classes = [permissions.IsAuthenticated & EstimatePermissions]
    filter_backends = (filters.DjangoFilterBackend, rf_filters.SearchFilter)
    filterset_class = DescriptionFilter
    search_fields = ('name', )


class DescriptionLibraryDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = DescriptionLibrary.objects.all()
    serializer_class = DescriptionLibrarySerializer
    permission_classes = [permissions.IsAuthenticated & EstimatePermissions]


class AssembleList(CompanyFilterMixin, generics.ListCreateAPIView):
    queryset = Assemble.objects.filter(
        estimate_templates=None, is_show=True
    ).order_by('-modified_date').prefetch_related(*ASSEMBLE_PREFETCH_RELATED)
    serializer_class = AssembleSerializer
    permission_classes = [permissions.IsAuthenticated & EstimatePermissions]
    filter_backends = (filters.DjangoFilterBackend, rf_filters.SearchFilter)
    filterset_class = AssembleFilter
    search_fields = ('name', 'description')


class AssembleCompactList(AssembleList):
    serializer_class = AssembleCompactSerializer


class AssembleDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = Assemble.objects.all().prefetch_related(*ASSEMBLE_PREFETCH_RELATED)
    serializer_class = AssembleSerializer
    permission_classes = [permissions.IsAuthenticated & EstimatePermissions]


class EstimateTemplateList(CompanyFilterMixin, generics.ListCreateAPIView):
    queryset = EstimateTemplate.objects.filter(
        is_show=True).order_by('-modified_date').prefetch_related(*ALL_ESTIMATE_PREFETCH_RELATED)
    serializer_class = EstimateTemplateSerializer
    permission_classes = [permissions.IsAuthenticated & EstimatePermissions]
    filter_backends = (filters.DjangoFilterBackend,)
    filterset_class = EstimateTemplateFilter


class EstimateTemplateCompactList(EstimateTemplateList):
    serializer_class = EstimateTemplateCompactSerializer


class EstimateTemplateDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = EstimateTemplate.objects.all().prefetch_related(*ALL_ESTIMATE_PREFETCH_RELATED)
    serializer_class = EstimateTemplateSerializer
    permission_classes = [permissions.IsAuthenticated & EstimatePermissions]


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated & EstimatePermissions])
def filter_group_fo_to_fo(request):
    q = Q()
    temp = request.query_params.dict()

    if temp == {}:
        grouping_queryset = POFormulaGrouping.objects.filter(
            company=request.user.company,
        ).order_by('-modified_date').distinct().values()
        for data in grouping_queryset:
            po = POFormula.objects.filter(Q(group_id=data['id']) & q)
            data['group_formulas'] = po
        paginator = LimitOffsetPagination()
        grouping_po_rs = paginator.paginate_queryset(grouping_queryset, request)
        serializer = POFormulaGroupingSerializer(grouping_po_rs, many=True)
        return paginator.get_paginated_response(serializer.data)
    data_filter = {
        "today": {
                "%s__year" % 'modified_date': now().year,
                "%s__month" % 'modified_date': now().month,
                "%s__day" % 'modified_date': now().day,
        },
        "yesterday": {
                "%s__gte" % 'modified_date': now() - timedelta(days=1),
                "%s__lt" % 'modified_date': now(),
        },
        "past-7-days": {
                "%s__gte" % 'modified_date': _truncate(now() - timedelta(days=7)),
                "%s__lt" % 'modified_date': _truncate(now() + timedelta(days=1)),
        },
        "past-14-days": {
                "%s__gte" % 'modified_date': _truncate(now() - timedelta(days=14)),
                "%s__lt" % 'modified_date': _truncate(now() + timedelta(days=1)),
        },
        "past-30-days": {
                "%s__gte" % 'modified_date': _truncate(now() - timedelta(days=30)),
                "%s__lt" % 'modified_date': _truncate(now() + timedelta(days=1)),
        },
        "past-45-days": {
                "%s__gte" % 'modified_date': _truncate(now() - timedelta(days=45)),
                "%s__lt" % 'modified_date': _truncate(now() + timedelta(days=1)),
        },
        "past-90-days": {
                "%s__gte" % 'modified_date': _truncate(now() - timedelta(days=90)),
                "%s__lt" % 'modified_date': _truncate(now() + timedelta(days=1)),
        },
    }
    q &= Q(**{f'company': request.user.company})
    for data in temp:
        if data == 'modified_date' and temp['modified_date'] != str():
            date = datetime.strptime(temp[data], '%Y-%m-%d %H:%M:%S')
            q &= Q(**{f'{data}__gt': date})
        if data == 'created_date' and temp['created_date'] != str():
            date = datetime.strptime(temp[data], '%Y-%m-%d %H:%M:%S')
            q &= Q(**{f'{data}__gt': date})
        if data == 'age_of_formula' and temp['age_of_formula'] != str():
            q &= Q(**data_filter[temp['age_of_formula']])
        if data != 'cost' and data != 'limit' and data != 'offset' and data != 'modified_date' and data != 'created_date' and data != 'age_of_formula':
            q &= Q(**{f'{data}__icontains': temp[data]})
        if data == 'cost' and temp[data] != str():
            q &= Q(cost__gt=temp[data])

    formula_queryset = POFormula.objects.filter(q)
    grouping_queryset = POFormulaGrouping.objects.filter(
        company=request.user.company,
        group_formulas__in=Subquery(formula_queryset.values('pk'))
    ).order_by('-modified_date').distinct().values()

    for data in grouping_queryset:
        po = POFormula.objects.filter(Q(group_id=data['id']) & q)
        data['group_formulas'] = po
    paginator = LimitOffsetPagination()
    grouping_po_rs = paginator.paginate_queryset(grouping_queryset, request)
    serializer = POFormulaGroupingSerializer(grouping_po_rs, many=True)
    return paginator.get_paginated_response(serializer.data)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated & EstimatePermissions])
def get_linked_descriptions(request):
    """
    Get linked description from estimate and catalog
    """
    search_query = {'linked_description__icontains': request.GET.get('search', '')}
    dl = DescriptionLibrary.objects.filter(**search_query, company=get_request().user.company)
    paginator = LimitOffsetPagination()
    estimate_result = paginator.paginate_queryset(dl, request)
    estimate_serializer = LinkedDescriptionSerializer(estimate_result, many=True)
    return paginator.get_paginated_response(estimate_serializer.data)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated & EstimatePermissions])
def get_linked_description(request, pk):
    """
    Get linked description from estimate and catalog
    """
    if 'estimate' in pk:
        obj = get_object_or_404(DescriptionLibrary.objects.filter(company=request.user.company), pk=pk.split(':')[1])
    else:
        obj = get_object_or_404(DataPoint.objects.filter(company=request.user.company), pk=pk.split(':')[1])
    serializer = LinkedDescriptionSerializer(obj)
    return Response(status=status.HTTP_200_OK, data=serializer.data)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated & EstimatePermissions])
def get_formula_tag_value(request, pk):
    """
    Get linked description from estimate and catalog
    """
    estimate = get_object_or_404(EstimateTemplate.objects.filter(company=request.user.company), pk=pk)
    formulas = estimate.get_formula()
    data = []
    for formula in formulas:
        data.extend(formula.parse_value_with_tag())
    return Response(status=status.HTTP_200_OK, data=data)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated & EstimatePermissions])
def get_tag_formula(request):
    formulas = POFormula.objects.filter(company=get_request().user.company, is_show=True)
    serializer = TaggingSerializer(formulas, many=True)
    return Response(data=serializer.data, status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated & EstimatePermissions])
def get_tag_data_point(request):
    # data_points = DataPoint.objects.filter(company=get_request().user.company)
    # serializer = TaggingSerializer(data_points, many=True)
    return Response(data=[], status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated & EstimatePermissions])
def get_tag_levels(request):
    material_id = request.GET.get('material_id')
    if not material_id:
        return Response(data='Missing material_id parameter!', status=status.HTTP_400_BAD_REQUEST)
    catalog_pk = material_id.split(':')[0]
    catalog = get_object_or_404(Catalog.objects.filter(company=request.user.company), pk=catalog_pk)
    ancestor = catalog.get_ancestors()

    serializer = CatalogLevelValueSerializer(ancestor[::-1], many=True)
    return Response(data=serializer.data, status=status.HTTP_200_OK)


@api_view(['PUT'])
@permission_classes([permissions.IsAuthenticated & EstimatePermissions])
def unlink_group(request):
    ids = request.data
    formulas = POFormula.objects.filter(id__in=ids)
    formulas.update(group=None)
    serializer = POFormulaSerializer(formulas, many=True)
    return Response(data=serializer.data, status=status.HTTP_200_OK)


@api_view(['PUT'])
@permission_classes([permissions.IsAuthenticated & EstimatePermissions])
def add_existing_formula(request, pk):
    ids = request.data
    formulas = POFormula.objects.filter(id__in=ids)
    formulas.update(group=pk)
    serializer = POFormulaSerializer(formulas, many=True)
    return Response(data=serializer.data, status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated & EstimatePermissions])
def get_material_by_data_entry(request, pk):
    de = DataEntry.objects.get(pk=pk)
    categories = de.material_selections.all()
    children = Catalog.objects.none()
    for category in categories:
        children |= Catalog.objects.filter(pk__in=category.get_all_descendant(have_self=True))
    children = children.difference(Catalog.objects.filter(c_table=Value('{}')))
    data = parse_c_table(children)
    return Response(status=status.HTTP_200_OK, data=data)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated & EstimatePermissions])
def get_material_from_formula(request, pk):
    mv = get_object_or_404(MaterialView.objects.filter(company=request.user.company), pk=pk)
    catagory_id = mv.catalog_materials[-1].get('id')
    cat = get_object_or_404(Catalog.objects.all(), pk=catagory_id)
    sub_categories = Catalog.objects.filter(pk__in=cat.get_all_descendant(have_self=True))
    data = parse_c_table(sub_categories)
    return Response(status=status.HTTP_200_OK, data=data)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated & EstimatePermissions])
def get_option_data_entry(request, pk):
    de = DataEntry.objects.get(pk=pk)
    categories = de.material_selections.all()
    children = Catalog.objects.none()
    for category in categories:
        children |= Catalog.objects.filter(parents=category)
    serializer = CatalogEstimateWithParentSerializer(children, many=True)
    return Response(status=status.HTTP_200_OK, data=serializer.data)


@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([permissions.IsAuthenticated & EstimatePermissions])
def check_multiple_formula_action(request):
    if request.method == 'GET':
        data = {'assembles': []}
        pk_formulas = request.GET.getlist('pk', [])
        formulas = POFormula.objects.filter(pk__in=pk_formulas)
        for po in formulas:
            data['assembles'].append({'formula': {'id': po.id, 'name': po.name},
                                       **po.get_related_formula()},)
        return Response(status=status.HTTP_200_OK, data=data)
    if request.method == 'DELETE':
        assembles_params = request.data.pop('assembles', [])
        for assemble in assembles_params:
            assembles_ids = assemble['assembles']
            pk = assemble['formula']['id']
            assembles = Assemble.objects.filter(id__in=assembles_ids)
            POFormula.objects.filter(original=pk, assemble__in=assembles, is_show=False).delete()
            POFormula.objects.filter(id=pk).delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


def clone_object(query, Serializer, request):
    new_data = {}
    for obj in query:
        data = Serializer(obj).data
        # data['name'] = data['name'] + f' {1}'
        serializer = Serializer(data=data, context={'request': request})
        serializer.is_valid()
        new_data[obj.id] = serializer.save(is_show=True, original=0)
    return new_data


def update_duplicate_name(Model, name):
    same_objs = Model.objects.filter(name__exact=name, is_show=True).order_by('created_date')
    has_duplicated_objs = Model.objects.filter(name__regex=rf'{name} (\d+)$')
    #  Get name of duplicated objects
    duplicated_names = [obj.name for obj in has_duplicated_objs]
    #  Get the number of duplicated objects
    duplicated_numbers = [int(re.search(rf'{name} (\d+)$', obj).group(1)) for obj in duplicated_names]
    for idx, obj in enumerate(same_objs):
        if idx:
            # get idx not in duplicated_numbers
            idx = next((i for i in range(1, 1000) if i not in duplicated_numbers), None)
            duplicated_numbers.append(idx)
            obj.name = f'{name} {idx}'
    Model.objects.bulk_update(same_objs, ['name'])


@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([permissions.IsAuthenticated & EstimatePermissions])
def action_related_formulas(request, pk):
    current_formula = get_object_or_404(POFormula.objects.all(), pk=pk)
    if request.method == 'GET':
        data = current_formula.get_related_formula()
        return Response(status=status.HTTP_200_OK, data=data)

    if request.method == 'PUT':
        formula_payload = request.data.pop('formula', {})
        assembles_params = request.data.pop('assembles', [])
        estimate_params = request.data.pop('estimates', [])
        writing_params = request.data.pop('writtings', [])
        comparison_params = request.data.pop('comparisons', [])
        delist_column = ['group', 'assemble', 'is_show', 'id', 'original', 'formula_for_data_view']
        for column in delist_column:
            formula_payload.pop(column, None)

        serializer = POFormulaSerializer(data=formula_payload)
        serializer.is_valid(raise_exception=True)
        new_obj = serializer.save(is_show=True)

        # update to proposal
        assembles = Assemble.objects.filter(id__in=assembles_params)
        new_assembles = clone_object(assembles, AssembleSerializer, request)

        formulas = POFormula.objects.filter(original=pk, assemble__in=new_assembles.values(), is_show=False)
        for formula in formulas:
            serializer = POFormulaSerializer(instance=formula, data=formula_payload, context={'request': request})
            serializer.is_valid(raise_exception=True)
            serializer.save(assemble=formula.assemble, original=new_obj.pk, is_show=False, group=None)

        # For estimate template
        estimates = EstimateTemplate.objects.filter(id__in=estimate_params)
        new_estimates = clone_object(estimates, EstimateTemplateSerializer, request)
        for estimate in new_estimates.values():
            estimate_assembles = Assemble.objects.filter(
                is_show=False, original__in=[obj.id for obj in assembles], estimate_templates=estimate
            )
            estimate_formulas = POFormula.objects.filter(assemble__in=estimate_assembles, original=pk)
            for formula in estimate_formulas:
                serializer = POFormulaSerializer(instance=formula, data=formula_payload, context={'request': request})
                serializer.is_valid(raise_exception=True)
                serializer.save(assemble=formula.assemble, original=new_obj.pk, is_show=False, group=None)
            change_assembles = []
            for assemble in estimate_assembles:
                assemble.original = new_assembles[assemble.original].pk
                change_assembles.append(assemble)
            Assemble.objects.bulk_update(change_assembles, ['original'])

        related_estimates = []
        for e in estimates:
            data_estimate = []
            writing_estimates = EstimateTemplate.objects.filter(original=e.pk, group_by_proposal__writing__id__in=writing_params)
            for estimate in writing_estimates:
                related_estimates.append(estimate)
                estimate_assembles = Assemble.objects.filter(
                    is_show=False, original__in=[obj.id for obj in assembles], estimate_templates=estimate
                )
                estimate_formulas = POFormula.objects.filter(assemble__in=estimate_assembles, original=pk)
                for formula in estimate_formulas:
                    serializer = POFormulaSerializer(instance=formula, data=formula_payload, context={'request': request})
                    serializer.is_valid(raise_exception=True)
                    serializer.save(assemble=formula.assemble, original=new_obj.pk, is_show=False, group=None)
                change_assembles = []
                for assemble in estimate_assembles:
                    assemble.original = new_assembles[assemble.original].pk
                    change_assembles.append(assemble)
                estimate.original = new_estimates[e.pk].pk
                data_estimate.append(estimate)
                Assemble.objects.bulk_update(change_assembles, ['original'])
            # Price comparison
            comparison_estimate = EstimateTemplate.objects.filter(original=estimate.pk, group_price__price_comparison__id__in=comparison_params)
            for estimate in comparison_estimate:
                related_estimates.append(estimate)
                estimate_assembles = Assemble.objects.filter(
                    is_show=False, original__in=[obj.id for obj in assembles], estimate_templates=estimate
                )
                estimate_formulas = POFormula.objects.filter(assemble__in=estimate_assembles, original=pk)
                for formula in estimate_formulas:
                    serializer = POFormulaSerializer(instance=formula, data=formula_payload, context={'request': request})
                    serializer.is_valid(raise_exception=True)
                    serializer.save(assemble=formula.assemble, original=new_obj.pk, is_show=False, group=None)
                change_assembles = []
                for assemble in estimate_assembles:
                    assemble.original = new_assembles[assemble.original].pk
                    change_assembles.append(assemble)
                estimate.original = new_estimates[e.pk].pk
                data_estimate.append(estimate)
                Assemble.objects.bulk_update(change_assembles, ['original'])
            EstimateTemplate.objects.bulk_update(data_estimate, ['original'])
        #  Clean data
        for e in estimates:
            name = e.name
            if not e.has_relation():
                e.delete()
            update_duplicate_name(EstimateTemplate, name)
        for assemble in assembles:
            name = assemble.name
            if not assemble.has_relation():
                assemble.delete()
            update_duplicate_name(Assemble, name)
        name = formula_payload['name']
        if not current_formula.has_relation():
            current_formula.delete()
        update_duplicate_name(POFormula, name)

        # sync new estimate
        for estimate in new_estimates.values():
            estimate.sync_data_entries()
        for estimate in related_estimates:
            estimate.sync_data_entries()
        return Response(status=status.HTTP_200_OK)

    if request.method == 'DELETE':
        assembles_ids = request.data.pop('assembles', [])
        assembles = Assemble.objects.filter(id__in=assembles_ids)
        POFormula.objects.filter(original=pk, assemble__in=assembles, is_show=False).delete()
        POFormula.objects.filter(id=pk).delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated & EstimatePermissions])
def export_data_unit_library(request):
    workbook = Workbook()
    unit_library_sheet = workbook.create_sheet(title='UnitLibrary')
    unit_library = UnitLibrary.objects.filter(company=request.user.company)
    unit_library_fields = ['Name', 'Description', 'User Create', 'User Update', 'Create Date', 'Update Date']
    unit_library_sheet.append(unit_library_fields)
    for index, data_unit_library in enumerate(unit_library, 1):
        user_create = data_unit_library.user_create.username
        user_update = data_unit_library.user_update
        create_date = data_unit_library.created_date
        if user_update is not None:
            user_update = user_update.username
        else:
            user_update = ''

        if create_date is not None:
            create_date = create_date.replace(tzinfo=None)
        else:
            create_date = ''

        modified_date = data_unit_library.modified_date
        if modified_date is not None:
            modified_date = modified_date.replace(tzinfo=None)
        else:
            modified_date = ''
        row_data = [
            data_unit_library.name, data_unit_library.description, user_create, user_update, create_date, modified_date
        ]

        unit_library_sheet.append(row_data)

    return file_response(workbook=workbook, title='Unit_Library')


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated & EstimatePermissions])
def import_data_unit_library(request):
    file = request.FILES['file']
    workbook = load_workbook(file)
    unit_library_sheet = workbook['UnitLibrary']

    max_row = unit_library_sheet.max_row
    temp_rs = []
    for row_number in range(max_row, 1, -1):
        row = unit_library_sheet[row_number]
        data_create = {
            'name': row[0].value,
            'description': row[1].value
        }
        ul = UnitLibrary.objects.create(**data_create)
        temp_rs.append(ul)

    rs = UnitLibrarySerializer(
        temp_rs, many=True, context={'request': request}).data
    return Response(status=status.HTTP_200_OK, data=rs)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated & EstimatePermissions])
def export_data_description_library(request):
    workbook = Workbook()
    description_library_sheet = workbook.create_sheet(title='LinkDescriptionLibrary')
    description_library = DescriptionLibrary.objects.filter(company=request.user.company)
    description_library_fields = ['Name', 'Link Description', 'User Create', 'User Update', 'Create Date', 'Update Date']
    description_library_sheet.append(description_library_fields)
    for index, data_description_library in enumerate(description_library, 1):
        user_create = data_description_library.user_create.username
        user_update = data_description_library.user_update
        create_date = data_description_library.created_date
        if user_update is not None:
            user_update = user_update.username
        else:
            user_update = ''

        if create_date is not None:
            create_date = create_date.replace(tzinfo=None)
        else:
            create_date = ''

        modified_date = data_description_library.modified_date
        if modified_date is not None:
            modified_date = modified_date.replace(tzinfo=None)
        else:
            modified_date = ''
        row_data = [
            data_description_library.name, data_description_library.linked_description, user_create, user_update, create_date, modified_date
        ]

        description_library_sheet.append(row_data)

    return file_response(workbook=workbook, title='Description_Library')


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated & EstimatePermissions])
def import_data_description_library(request):
    file = request.FILES['file']
    workbook = load_workbook(file)
    description_library_sheet = workbook['LinkDescriptionLibrary']

    max_row = description_library_sheet.max_row
    temp_rs = []
    for row_number in range(max_row, 1, -1):
        row = description_library_sheet[row_number]
        data_create = {
            'name': row[0].value,
            'linked_description': row[1].value
        }
        ul = DescriptionLibrary.objects.create(**data_create)
        temp_rs.append(ul)

    rs = DescriptionLibrarySerializer(
        temp_rs, many=True, context={'request': request}).data
    return Response(status=status.HTTP_200_OK, data=rs)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated & EstimatePermissions])
def export_data_entry(request):
    workbook = Workbook()
    data_entry_sheet = workbook.create_sheet(title='Data_Entry')
    data_entries = DataEntry.objects.filter(company=request.user.company)
    data_entry_fields = ['Name', 'Unit', 'Is Dropdown', 'Dropdown', 'Is Material Selection', 'Material Selection']
    data_entry_sheet.append(data_entry_fields)
    for data_entry in data_entries:
        data_entry_sheet.append(data_entry.export_to_json())
    return file_response(workbook=workbook, title='Data Entry')


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated & EstimatePermissions])
def import_data_entry(request):
    file = request.FILES['file']
    workbook = load_workbook(file)
    data_entry_sheet = workbook['Data_Entry']

    max_row = data_entry_sheet.max_row
    temp_rs = []
    for row_number in range(max_row, 1, -1):
        row = data_entry_sheet[row_number]
        unit = None
        if row[1]:
            unit = UnitLibrary.objects.get_or_create(name=row[1].value, company=request.user.company)[0]
        data_create = {
            'name': row[0].value,
            'unit': unit,
            'is_dropdown': row[2].value,
            'dropdown': eval(row[3].value),
            'is_material_selection': row[4].value,
        }

        ul = DataEntry.objects.create(**data_create)
        if row[5].value:
            ul.material_selections.add(*Catalog.objects.filter(id__in=row[5].value))
        temp_rs.append(ul)

    rs = DataEntrySerializer(
        temp_rs, many=True, context={'request': request}).data
    return Response(status=status.HTTP_200_OK, data=rs)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated & EstimatePermissions])
def export_formula(request):
    workbook = Workbook()
    formula_sheet = workbook.create_sheet(title='Formula')
    formulas = POFormula.objects.filter(company=request.user.company, is_show=True)
    formula_fields = ['Name', 'Linked Description', 'Formula', 'Group', 'Quantity', 'Markup', 'Charge',
                      'Material', 'Unit', 'Unit Price', 'Cost', 'Total Cost', 'Formula Mention', 'Gross Profit',
                      'Description', 'Scenario', 'Material Data Entry', 'Catalog Material']
    formula_sheet.append(formula_fields)
    for formula in formulas:
        formula_sheet.append(formula.export_to_json())
    return file_response(workbook=workbook, title='Formula')


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated & EstimatePermissions])
def import_formula(request):
    file = request.FILES['file']
    workbook = load_workbook(file)
    formula_sheet = workbook['Formula']

    max_row = formula_sheet.max_row
    temp_rs = []
    for row_number in range(max_row, 1, -1):
        row = formula_sheet[row_number]
        data_create = {
            'name': row[0].value,
            'linked_description': row[1].value or '',
            'formula': row[2].value or '',
            'quantity': row[4].value or '',
            'markup': row[5].value or '',
            'charge': row[6].value or '0',
            'material': row[7].value or {},
            'unit': row[8].value or '',
            'unit_price': row[9].value or '0',
            'cost': row[10].value or '0',
            'total_cost': row[11].value or '0',
            'formula_mentions': row[12].value or '',
            'gross_profit': row[13].value or '',
            'description_of_formula': row[14].value or '',
            'formula_scenario': row[15].value or '',
            'material_data_entry': eval(row[16].value),
            'catalog_materials': eval(row[17].value),
        }

        ul = POFormula.objects.create(**data_create)
        temp_rs.append(ul)

    rs = POFormulaGroupCompactSerializer(
        temp_rs, many=True, context={'request': request}).data
    return Response(status=status.HTTP_200_OK, data=rs)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated & EstimatePermissions])
def get_formula_description(request, pk):
    formula = get_object_or_404(POFormula.objects.all(), pk=pk)
    catalog_params = request.GET.getlist('catalogs', [])
    catalogs = Catalog.objects.filter(pk__in=catalog_params)
    data = {}
    data['formula'] = formula.linked_description
    data['catalogs'] = CatalogSerializer(catalogs, many=True).data
    return Response(status=status.HTTP_200_OK, data=data)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated & EstimatePermissions])
def export_assemble(request):
    workbook = Workbook()
    assemble_sheet = workbook.create_sheet(title='Assemble')
    assembles = Assemble.objects.filter(company=request.user.company, is_show=True)
    assemble_fields = ['Name', 'Meta']
    assemble_sheet.append(assemble_fields)
    for assemble in assembles:
        assemble_sheet.append([*assemble.export_to_json(), str(json.dumps(AssembleSerializer(assemble).data))])
    return file_response(workbook=workbook, title='Assemble')


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated & EstimatePermissions])
def export_estimate(request):
    workbook = Workbook()
    estimate_sheet = workbook.create_sheet(title='Estimate')
    estimates = EstimateTemplate.objects.filter(company=request.user.company, is_show=True)
    estimate_fields = ['Name', 'Meta']
    estimate_sheet.append(estimate_fields)
    for estimate in estimates:
        estimate_sheet.append([*estimate.export_to_json(), str(json.dumps(EstimateTemplateSerializer(estimate).data))])
    return file_response(workbook=workbook, title='Estimate')


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated & EstimatePermissions])
def import_assemble(request):
    file = request.FILES['file']
    workbook = load_workbook(file)
    assemble_sheet = workbook['Assemble']

    max_row = assemble_sheet.max_row
    temp_rs = []
    for row_number in range(max_row, 1, -1):
        row = assemble_sheet[row_number]
        data_create = {
            'name': row[0].value,
            **eval(row[1].value)
        }

        serializer = AssembleSerializer(data=data_create, context={'request': request})
        if serializer.is_valid(raise_exception=False):
            temp_rs.append(serializer.save())

    rs = AssembleSerializer(
        temp_rs, many=True, context={'request': request}).data
    return Response(status=status.HTTP_200_OK, data=rs)


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated & EstimatePermissions])
def import_estimate(request):
    file = request.FILES['file']
    workbook = load_workbook(file)
    estimate_sheet = workbook['Estimate']

    max_row = estimate_sheet.max_row
    temp_rs = []
    for row_number in range(max_row, 1, -1):
        row = estimate_sheet[row_number]
        data_create = {
            'name': row[0].value,
            **eval(row[1].value)
        }
        serializer = EstimateTemplateSerializer(data=data_create, context={'request': request})
        if serializer.is_valid(raise_exception=False):
            temp_rs.append(serializer.save())

    rs = EstimateTemplateSerializer(
        temp_rs, many=True, context={'request': request}).data
    return Response(status=status.HTTP_200_OK, data=rs)


@api_view(['GET', 'PUT'])
@permission_classes([permissions.IsAuthenticated & EstimatePermissions])
def check_update_data_entry(request, pk):
    """
    Payload: {"data_entry": Data Entry object, "formulas": [id]}

    Updating Data Entry for formulas
    """
    data_entry = get_object_or_404(DataEntry.objects.all(), pk=pk)
    self_data_entries = data_entry.poformulatodataentry_set.all()
    if request.method == 'GET':
        data = {}
        data['has_relation'] = data_entry.poformulatodataentry_set.all().exists()
        formulas = POFormula.objects.filter(self_data_entries__in=self_data_entries, is_show=True).distinct()
        data['formula_relation'] = formulas.values('id', 'name')
        data['data_entry'] = DataEntrySerializer(data_entry).data
        return Response(status=status.HTTP_200_OK, data=data)

    if request.method == 'PUT':
        data_entry_params = request.data.get('data_entry', {})
        formula_params = request.data.get('formulas', [])
        assembles_params = request.data.pop('assembles', [])
        estimate_params = request.data.pop('estimates', [])
        writing_params = request.data.pop('writtings', [])
        comparison_params = request.data.pop('comparisons', [])
        data_entry_params.pop('id', None)
        new_serializer = DataEntrySerializer(data=data_entry_params)
        new_serializer.is_valid(raise_exception=True)
        new_obj = new_serializer.save()

        formulas = POFormula.objects.filter(pk__in=formula_params, is_show=True)
        new_formulas = clone_object(formulas, POFormulaSerializer, request)
        formula_with_data_entry = POFormulaToDataEntry.objects.filter(data_entry=data_entry, po_formula__in=new_formulas.values())
        update_po_data_entry(formula_with_data_entry, new_obj, data_entry)

        # update to proposal
        assembles = Assemble.objects.filter(id__in=assembles_params)
        new_assembles = clone_object(assembles, AssembleSerializer, request)

        for old_formula in formulas:
            pos = POFormula.objects.filter(original=old_formula.pk, assemble__in=new_assembles.values(), is_show=False)
            formula_with_data_entry = POFormulaToDataEntry.objects.filter(data_entry=data_entry, po_formula__in=pos)
            update_po_data_entry(formula_with_data_entry, new_obj, data_entry)
            pos.update(original=new_formulas[old_formula.pk].pk)

        estimates = EstimateTemplate.objects.filter(id__in=estimate_params)
        new_estimates = clone_object(estimates, EstimateTemplateSerializer, request)
        for estimate in new_estimates.values():
            estimate_assembles = Assemble.objects.filter(
                is_show=False, original__in=[obj.id for obj in assembles], estimate_templates=estimate
            )
            estimate_formulas = POFormula.objects.filter(assemble__in=estimate_assembles, original__in=[obj.id for obj in formulas])
            for formula in estimate_formulas:
                # Update original formula
                formula.original = new_formulas[formula.original].pk
                formula_with_data_entry = POFormulaToDataEntry.objects.filter(data_entry=data_entry, po_formula__in=estimate_formulas)
                update_po_data_entry(formula_with_data_entry, new_obj, data_entry)
            POFormula.objects.bulk_update(estimate_formulas, ['original'])
            change_assembles = []
            for assemble in estimate_assembles:
                assemble.original = new_assembles[assemble.original].pk
                change_assembles.append(assemble)
            Assemble.objects.bulk_update(change_assembles, ['original'])

        related_estimates = []
        for e in estimates:
            data_estimate = []
            writing_estimates = EstimateTemplate.objects.filter(original=e.pk, group_by_proposal__writing__id__in=writing_params)
            for estimate in writing_estimates:
                related_estimates.append(estimate)
                estimate_assembles = Assemble.objects.filter(
                    is_show=False, original__in=[obj.id for obj in assembles], estimate_templates=estimate
                )
                estimate_formulas = POFormula.objects.filter(assemble__in=estimate_assembles, original__in=[obj.id for obj in formulas])
                for formula in estimate_formulas:
                    # Update original formula
                    formula.original = new_formulas[formula.original].pk
                    formula_with_data_entry = POFormulaToDataEntry.objects.filter(data_entry=data_entry, po_formula__in=estimate_formulas)
                    update_po_data_entry(formula_with_data_entry, new_obj, data_entry)
                POFormula.objects.bulk_update(estimate_formulas, ['original'])
                change_assembles = []
                for assemble in estimate_assembles:
                    assemble.original = new_assembles[assemble.original].pk
                    change_assembles.append(assemble)
                estimate.original = new_estimates[e.pk].pk
                data_estimate.append(estimate)
                Assemble.objects.bulk_update(change_assembles, ['original'])
            # Price comparison
            comparison_estimate = EstimateTemplate.objects.filter(original=estimate.pk, group_price__price_comparison__id__in=comparison_params)
            for estimate in comparison_estimate:
                related_estimates.append(estimate)
                estimate_assembles = Assemble.objects.filter(
                    is_show=False, original__in=[obj.id for obj in assembles], estimate_templates=estimate
                )
                estimate_formulas = POFormula.objects.filter(assemble__in=estimate_assembles, original=[obj.id for obj in formulas])
                for formula in estimate_formulas:
                    # Update original formula
                    formula.original = new_formulas[formula.original].pk
                    formula_with_data_entry = POFormulaToDataEntry.objects.filter(data_entry=data_entry, po_formula__in=estimate_formulas)
                    update_po_data_entry(formula_with_data_entry, new_obj, data_entry)
                    POFormula.objects.bulk_update(estimate_formulas, ['original'])
                change_assembles = []
                for assemble in estimate_assembles:
                    assemble.original = new_assembles[assemble.original].pk
                    change_assembles.append(assemble)
                estimate.original = new_estimates[e.pk].pk
                data_estimate.append(estimate)
                Assemble.objects.bulk_update(change_assembles, ['original'])
            EstimateTemplate.objects.bulk_update(data_estimate, ['original'])

        for e in estimates:
            name = e.name
            if not e.has_relation():
                e.delete()
            update_duplicate_name(EstimateTemplate, name)
        for assemble in assembles:
            name = assemble.name
            if not assemble.has_relation():
                assemble.delete()
            update_duplicate_name(Assemble, name)
        for formula in formulas:
            name = formula.name
            if not formula.has_relation():
                formula.delete()
            update_duplicate_name(POFormula, name)
        name = data_entry_params['name']
        if not data_entry.has_relation():
            data_entry.delete()
        update_duplicate_name(DataEntry, name)
        # sync new estimate
        for estimate in new_estimates.values():
            estimate.sync_data_entries()
        for estimate in related_estimates:
            estimate.sync_data_entries()
        for proposal in ProposalWriting.objects.filter(id__in=writing_params):
            proposal.update_info()
        return Response(status=status.HTTP_200_OK, data={'data_entry': data_entry_params})


def update_po_data_entry(formula_with_data_entry, new_obj, old_obj):
    data = []
    for po in formula_with_data_entry:
        po = po.po_formula
        if hasattr(po, 'formula_mentions'):
            po.formula_mentions = po.formula_mentions.replace(f'$[{old_obj.name}]({old_obj.pk})', f'$[{new_obj.name}]({new_obj.pk})')
            data.append(po)
    formula_with_data_entry.update(data_entry=new_obj)
    POFormula.objects.bulk_update(data, ['formula_mentions'])


@api_view(['GET', 'PUT'])
@permission_classes([permissions.IsAuthenticated & EstimatePermissions])
def check_update_estimate(request, pk):
    ProposalWriting = apps.get_model('sales', 'ProposalWriting')
    PriceComparison = apps.get_model('sales', 'PriceComparison')
    estimate = get_object_or_404(EstimateTemplate.objects.all(), pk=pk)
    proposal_writings = ProposalWriting.objects.filter(writing_groups__estimate_templates__original=pk).distinct()
    price_comparisons = PriceComparison.objects.filter(groups__estimate_templates__original=pk).distinct()

    if request.method == 'PUT':
        estimate_params = request.data.get('estimate', {})
        proposal_writing_params = request.data.get('writtings', [])
        price_comparison_params = request.data.get('comparisons', [])
        # Get related estimate on proposal writings
        writing_estimates = EstimateTemplate.objects.filter(original=pk, group_by_proposal__writing__id__in=proposal_writing_params)

        for e in writing_estimates:
            serializer = EstimateTemplateSerializer(data=estimate_params, instance=e, context={'request': request})
            serializer.is_valid()
            serializer.save(group_by_proposal=e.group_by_proposal, is_show=False, original=pk)

        # Price comparison
        comparison_estimate = EstimateTemplate.objects.filter(original=pk, group_price__price_comparison__id__in=price_comparison_params)
        for e in comparison_estimate:
            serializer = EstimateTemplateSerializer(data=estimate_params, instance=e, context={'request': request})
            serializer.is_valid()
            serializer.save(is_show=False, original=pk)

        # Get proposal
        proposal_writings = ProposalWriting.objects.filter(id__in=proposal_writing_params)
        for proposal in proposal_writings:
            proposal.update_info()
        # Estimate template
        serializer = EstimateTemplateSerializer(data=estimate_params, instance=estimate, context={'request': request})
        serializer.is_valid()
        serializer.save(is_show=True, original=pk)
        return Response(status=status.HTTP_200_OK, data=serializer.data)
    data = {}
    data['has_relation'] = proposal_writings.exists() or price_comparisons.exists()
    data['writtings'] = proposal_writings.values('id', 'name')
    data['comparisons'] = price_comparisons.values('id', 'name')

    return Response(status=status.HTTP_200_OK, data=data)


@api_view(['GET', 'PUT'])
@permission_classes([permissions.IsAuthenticated & EstimatePermissions])
def check_action_assemble(request, pk):
    Assemble = apps.get_model('sales', 'Assemble')
    assemble = get_object_or_404(Assemble.objects.all(), pk=pk)
    if request.method == 'GET':
        return Response(status=status.HTTP_200_OK)
    if request.method == 'PUT':
        assembles_param = request.data.pop('assemble', {})
        estimate_params = request.data.pop('estimates', [])
        writing_params = request.data.pop('writtings', [])
        comparison_params = request.data.pop('comparisons', [])
        # Create new assemble using serializer
        serializer = AssembleSerializer(data=assembles_param, context={'request': request})
        serializer.is_valid(raise_exception=True)
        new_assemble = serializer.save(is_show=True)

        # Update to estimate template
        estimates = EstimateTemplate.objects.filter(id__in=estimate_params)
        new_estimates = clone_object(estimates, EstimateTemplateSerializer, request)
        for estimate in new_estimates.values():
            estimate_assembles = Assemble.objects.filter(
                is_show=False, original=pk, estimate_templates=estimate
            )
            for obj in estimate_assembles:
                serializer = AssembleSerializer(instance=obj, data=assembles_param, context={'request': request})
                serializer.is_valid(raise_exception=True)
                serializer.save(is_show=False, original=new_assemble.pk)

        related_estimates = []
        for e in estimates:
            data_estimate = []
            writing_estimates = EstimateTemplate.objects.filter(original=e.pk, group_by_proposal__writing__id__in=writing_params)
            for estimate in writing_estimates:
                related_estimates.append(estimate)
                estimate_assembles = Assemble.objects.filter(
                    is_show=False, original=assemble.pk, estimate_templates=estimate
                )
                for obj in estimate_assembles:
                    serializer = AssembleSerializer(instance=obj, data=assembles_param, context={'request': request})
                    serializer.is_valid(raise_exception=True)
                    serializer.save(is_show=False, original=new_assemble.pk)
                estimate.original = new_estimates[e.pk].pk
                data_estimate.append(estimate)
            # Price comparison
            comparison_estimate = EstimateTemplate.objects.filter(original=estimate.pk, group_price__price_comparison__id__in=comparison_params)
            for estimate in comparison_estimate:
                related_estimates.append(estimate)
                estimate_assembles = Assemble.objects.filter(
                    is_show=False, original=assemble.pk, estimate_templates=estimate
                )
                for obj in estimate_assembles:
                    serializer = AssembleSerializer(instance=obj, data=assembles_param, context={'request': request})
                    serializer.is_valid(raise_exception=True)
                    serializer.save(is_show=False, original=new_assemble.pk)
                estimate.original = new_estimates[e.pk].pk
                data_estimate.append(estimate)
            EstimateTemplate.objects.bulk_update(data_estimate, ['original'])
        # clean old assemble
        for e in estimates:
            name = e.name
            if not e.has_relation():
                e.delete()
            update_duplicate_name(EstimateTemplate, name)
        name = assemble.name
        if not assemble.has_relation():
            assemble.delete()
        update_duplicate_name(Assemble, name)
        # sync new estimate
        for estimate in new_estimates.values():
            estimate.sync_data_entries()
        for estimate in related_estimates:
            estimate.sync_data_entries()
        for proposal in ProposalWriting.objects.filter(id__in=writing_params):
            proposal.update_info()
        return Response(status=status.HTTP_200_OK)


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated & EstimatePermissions])
def duplicate_formula(request):
    formulas = request.data
    for formula in POFormula.objects.filter(id__in=formulas):
        data = POFormulaSerializer(formula, context={'request': request}).data
        serializer = POFormulaSerializer(data=data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        serializer.save(is_show=True, original=0, group=None)
        update_duplicate_name(POFormula, formula.name)
    return Response(status=status.HTTP_200_OK)
