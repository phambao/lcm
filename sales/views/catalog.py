from celery.result import AsyncResult
from django.core.exceptions import ObjectDoesNotExist
from django.db import transaction
from django.db.models import Value
from django.shortcuts import get_object_or_404
from django_filters import rest_framework as filters
from django.apps import apps
from openpyxl.workbook import Workbook
from rest_framework import generics, permissions, status, filters as rf_filters
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response

from base.constants import URL_CLOUD
from base.models.config import FileBuilder365
from base.permissions import CatalogPermissions
from base.serializers.base import FileBuilder365ResSerializer
from base.tasks import import_catalog_task, process_export_catalog
from base.utils import file_response
from ..filters.catalog import CatalogFilter
from ..models.catalog import Catalog, CatalogLevel, DataPointUnit, CostTableTemplate
from ..serializers import catalog
from ..serializers.catalog import CatalogEstimateSerializer, CatalogSerializer
from api.middleware import get_request
from base.views.base import CompanyFilterMixin

UnitLibrary = apps.get_model(app_label='sales', model_name='UnitLibrary')


class CatalogList(CompanyFilterMixin, generics.ListCreateAPIView):
    queryset = Catalog.objects.all().prefetch_related('data_points', 'parents', 'children')
    serializer_class = catalog.CatalogSerializer
    permission_classes = [permissions.IsAuthenticated & CatalogPermissions]
    filter_backends = (filters.DjangoFilterBackend, rf_filters.SearchFilter)
    filterset_class = CatalogFilter
    search_fields = ('name',)


class CatalogDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = Catalog.objects.all().prefetch_related('data_points', 'parents', 'children')
    serializer_class = catalog.CatalogSerializer
    permission_classes = [permissions.IsAuthenticated & CatalogPermissions]

    def delete(self, request, *args, **kwargs):
        pk_catalog = kwargs.get('pk')
        data_catalog = catalog.Catalog.objects.filter(id=pk_catalog).first()
        c_table = data_catalog.c_table
        parent_catalog = data_catalog.parents.first()
        if parent_catalog:
            children_catalog = catalog.Catalog.objects.filter(parents=parent_catalog.id).count()
            if children_catalog <= 1:
                parent_catalog.c_table = c_table
                parent_catalog.save()

        return super().delete(request, *args, **kwargs)


class CatalogLevelList(generics.ListCreateAPIView):
    queryset = CatalogLevel.objects.all()
    serializer_class = catalog.CatalogLevelModelSerializer
    permission_classes = [permissions.IsAuthenticated & CatalogPermissions]

    def get_queryset(self):
        if getattr(self, 'swagger_fake_view', False):
            return CatalogLevel.objects.none()
        catalog = get_object_or_404(Catalog.objects.all(), pk=self.kwargs['pk_catalog'])
        try:
            ancester_level = catalog.all_levels.get(parent=None)
        except ObjectDoesNotExist:
            return []
        return ancester_level.get_ordered_descendant()

    def perform_create(self, serializer):
        instance = serializer.save()
        instance.catalog_id = self.kwargs['pk_catalog']
        instance.save()


class CatalogLevelDetail(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = catalog.CatalogLevelModelSerializer
    permission_classes = [permissions.IsAuthenticated & CatalogPermissions]

    def get_queryset(self):
        try:
            catalog = get_object_or_404(Catalog.objects.all(), pk=self.kwargs['pk_catalog'])
        except KeyError:
            # When swagger call this view, that doesn't pass param pk_catalog
            return CatalogLevel.objects.all()
        return catalog.all_levels.all()


class DataPointUnitView(CompanyFilterMixin, generics.ListCreateAPIView):
    serializer_class = catalog.DataPointUnitSerializer
    permission_classes = [permissions.IsAuthenticated & CatalogPermissions]
    queryset = DataPointUnit.objects.all()


class DataPointUnitDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = catalog.DataPointUnitSerializer
    permission_classes = [permissions.IsAuthenticated & CatalogPermissions]
    queryset = DataPointUnit.objects.all()


class CostTableTemplateListView(CompanyFilterMixin, generics.ListCreateAPIView):
    serializer_class = catalog.CostTableTemplateSerializer
    permission_classes = [permissions.IsAuthenticated & CatalogPermissions]
    queryset = CostTableTemplate.objects.all()
    filter_backends = (rf_filters.SearchFilter,)
    search_fields = ('name',)


class CostTableTemplateDetailView(CompanyFilterMixin, generics.RetrieveUpdateDestroyAPIView):
    serializer_class = catalog.CostTableTemplateSerializer
    permission_classes = [permissions.IsAuthenticated & CatalogPermissions]
    queryset = CostTableTemplate.objects.all()


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated & CatalogPermissions])
def get_catalog_children(request, pk):
    level = None
    catalogs = Catalog.objects.filter(parents__id=pk)
    for c in catalogs:
        if c.level:
            level = c.level
            break
    if level:
        catalogs = Catalog.objects.filter(parents__id=pk, level=level)

    serializer = catalog.CatalogSerializer(catalogs, many=True, context={'request': request})
    return Response(status=status.HTTP_200_OK, data=serializer.data)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated & CatalogPermissions])
def get_catalog_levels(request, pk):
    level = None
    catalogs = Catalog.objects.filter(parents__id=pk)
    for c in catalogs:
        if c.level:
            level = c.level
            break

    if not level:
        return Response(status=status.HTTP_200_OK, data=[])

    catalog_level = CatalogLevel.objects.get(pk=level.pk)
    catalog_levels = catalog_level.get_ordered_descendant()
    serializer = catalog.CatalogLevelModelSerializer(catalog_levels, many=True)
    return Response(status=status.HTTP_200_OK, data=serializer.data)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated & CatalogPermissions])
def get_catalog_tree(request, pk):
    catalog = get_object_or_404(Catalog, pk=pk)
    catalog_tree = catalog.get_tree_view()
    return Response(status=status.HTTP_200_OK, data=catalog_tree)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated & CatalogPermissions])
def get_catalog_list(request, pk):
    catalog_obj = get_object_or_404(Catalog, pk=pk)
    catalog_ids = catalog_obj.get_all_descendant()
    catalogs = Catalog.objects.filter(pk__in=catalog_ids).prefetch_related('data_points', 'parents', 'data_points__unit', 'children')
    serializer = catalog.CatalogSerializer(catalogs, many=True, context={'request': request})
    return Response(status=status.HTTP_200_OK, data=serializer.data)


@api_view(['DELETE'])
@permission_classes([permissions.IsAuthenticated & CatalogPermissions])
def delete_catalogs(request):
    ids = request.data
    catalogs = Catalog.objects.filter(pk__in=ids)
    for catalog in catalogs:
        catalog.delete()
    return Response(status=status.HTTP_204_NO_CONTENT)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated & CatalogPermissions])
def get_catalog_ancestors(request):
    ids = request.GET.getlist('id', [])
    data = {}
    if ids:
        catalogs = Catalog.objects.filter(id__in=ids, company=get_request().user.company).prefetch_related('parents')
        for c in catalogs:
            navigation = c.get_ancestors()
            navigation = navigation[1:]
            data[c.pk] = [n.id for n in navigation[::-1]]
    return Response(status=status.HTTP_200_OK, data=data)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated & CatalogPermissions])
def get_datapoint_by_catalog(request, pk):
    c = Catalog.objects.get(id=pk)
    serializer = catalog.DataPointSerializer(c.get_ancestor_linked_description(),
                                             many=True, context={'request': request})
    return Response(status=status.HTTP_200_OK, data=serializer.data)


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated & CatalogPermissions])
def add_multiple_level(request):
    if isinstance(request.data, dict):
        catalog_serializer = catalog.CatalogSerializer(data=request.data.get('catalog'), context={'request': request})
        catalog_serializer.is_valid(raise_exception=True)
        c = catalog_serializer.save()
        parent = None
        data = []
        for name in request.data.get('levels', []):
            if name:  # in case of empty string
                catalog_level = CatalogLevel.objects.create(name=name, parent=parent, catalog=c)
                parent = catalog_level
                data.append(catalog_level)
        serializer = catalog.CatalogLevelModelSerializer(data, many=True, context={'request': request})
        return Response(status=status.HTTP_201_CREATED, data={"levels": serializer.data,
                                                              "catalog": catalog_serializer.data})
    return Response(status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated & CatalogPermissions])
def duplicate_catalogs(request, pk):
    """
    Payload: [{"id": int, "depth": int, data_points: [id,...], descendant: [id,...]},...]
    """

    parent_catalog = get_object_or_404(Catalog, pk=pk)
    new_c = []
    data = []
    # duplicate by level
    if isinstance(request.data, list):
        for d in request.data:
            depth = int(d.get('depth', 0))
            data_points = d.get('data_points', [])
            try:
                c = Catalog.objects.get(pk=d.get('id'))
                new_c.append(c.duplicate(parent=parent_catalog, depth=depth, data_points=data_points))
            except Catalog.DoesNotExist:
                pass
        for new in new_c:
            data.extend(new.get_all_descendant(have_self=True))
        return Response(status=status.HTTP_201_CREATED,
                        data=catalog.CatalogSerializer(Catalog.objects.filter(id__in=data),
                                                       many=True, context={'request': request}).data)
    return Response(status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated & CatalogPermissions])
def duplicate_catalogs_on_tree(request, pk):
    """
    Payload: {data_points: [id,...], descendant: [id,...], level: int}
    """
    parent_catalog = get_object_or_404(Catalog, pk=pk)
    if isinstance(request.data, dict):
        data_points = request.data.get('data_points', [])
        descendant = request.data.get('descendant', [])
        level = request.data.get('level', None)
        parents = Catalog.objects.filter(pk__in=descendant, level_id=level)
        for root in parents:
            try:
                root.duplicate_by_catalog(parent=parent_catalog, descendant=descendant,
                                          data_points=data_points)
            except Catalog.DoesNotExist:
                pass
        data = Catalog.objects.filter(pk__in=parent_catalog.get_all_descendant())
        serializer = catalog.CatalogSerializer(data, many=True, context={'request': request})
        return Response(status=status.HTTP_201_CREATED, data=serializer.data)
    return Response(status=status.HTTP_400_BAD_REQUEST)


@api_view(['PUT'])
@permission_classes([permissions.IsAuthenticated & CatalogPermissions])
@transaction.atomic
def swap_level(request, pk_catalog):
    """
    Payload: {"levels": [{"id": int, "parent": int},...], "catalogs": {"parent_id": [children_id,...],...}}
    """
    if isinstance(request.data, dict):
        ancestor_catalog = get_object_or_404(Catalog, pk=pk_catalog)
        try:
            levels = request.data.get('levels')
            catalogs = request.data.get('catalogs')

            updated_level = []
            level_id = []
            for l in levels:
                pk = l.get('id')
                level_id.append(pk)
                level = CatalogLevel.objects.get(pk=pk)
                level.parent_id = l.get('parent')
                updated_level.append(level)
            ancestor_catalog.all_levels.filter(pk__in=level_id).update(parent=None)
            CatalogLevel.objects.bulk_update(updated_level, fields=['parent'])

            # Validate level
            ordered_level = ancestor_catalog.get_ordered_levels()

            updated_catalog = []
            for key in catalogs.keys():
                parent_catalog = Catalog.objects.get(pk=key)
                children_catalog = Catalog.objects.filter(id__in=catalogs[key])
                for c in children_catalog:
                    c.parents.clear()
                    c.parents.add(parent_catalog)
                    c.level_index = ordered_level.index(c.level)
                    updated_catalog.append(c)

            # Validate catalog
            if Catalog.objects.filter(level__in=level_id).count() != len(updated_catalog):
                raise Exception('Missing or too much categories updated')

            # Update level index for all categories
            for l in ordered_level:
                Catalog.objects.filter(level=l).update(level_index=ordered_level.index(l))

            catalog_serializer = catalog.CatalogSerializer(updated_catalog, many=True,
                                                           context={'request': request})
            level_serializer = catalog.CatalogLevelModelSerializer(updated_level, many=True,
                                                                   context={'request': request})

            return Response(status=status.HTTP_200_OK, data={
                "levels": level_serializer.data,
                "catalog": catalog_serializer.data
            })
        except Exception as e:
            transaction.set_rollback(True)
            raise e
    return Response(status=status.HTTP_400_BAD_REQUEST)


def parse_dict(dictionary, is_formula=[], actual_data=[]):
    data = []
    i = 0
    for name, value in dictionary.items():
        data.append({'name': name, 'value': value, 'is_formula': is_formula[i] if i < len(is_formula) else False, 
                     'default_value': actual_data[i] if i < len(actual_data) else 0})
        i += 1
    return data


def parse_c_table(children):
    data = []
    for child in children:
        try:
            try:
                ancestor = child.get_full_ancestor()
                levels = [CatalogEstimateSerializer(c).data for c in ancestor[::-1]]
            except:
                levels = []
            c_table = child.c_table
            header = c_table['header']

            if len(header) >= 3:
                header[:3] = 'name', 'unit', 'cost'
            header_formats = [is_formula.get('isFormula') for is_formula in c_table.get('header_format', [])]
            actual_data = c_table.get('actual_data', [])
            for i, d in enumerate(c_table['data']):
                actual_value = actual_data[i] if i < len(actual_data) else []
                content = {**{header[j]: d[j] for j in range(len(header))}, **{"id": f'{child.pk}:{i}'},
                           'levels': levels}
                clone = content.copy()
                clone.pop('id')
                clone.pop('levels')
                content['columns'] = parse_dict(clone, header_formats, actual_value)
                data.append(content)
        except:
            """Some old data is not valid"""
    return data


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated & CatalogPermissions])
def get_materials(request):
    """
    Get cost table from ancestor catalog
    """
    filter_query = request.GET.get('catalog', None)
    if filter_query:
        c = get_object_or_404(Catalog.objects.all(), pk=filter_query)
        children = Catalog.objects.filter(
            pk__in=c.get_all_descendant(have_self=True)
        )
    else:
        children = Catalog.objects.filter(company=get_request().user.company)
    children = children.difference(Catalog.objects.filter(c_table=Value('{}')))
    data = parse_c_table(children)
    return Response(status=status.HTTP_200_OK, data=data)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated & CatalogPermissions])
def get_all_cost_table(request):
    """
    Get cost table from ancestor catalog
    """
    filter_query = request.GET.get('catalog', None)
    if filter_query:
        c = get_object_or_404(Catalog.objects.all(), pk=filter_query)
        children = Catalog.objects.filter(
            pk__in=c.get_all_descendant(have_self=True)
        )
    else:
        children = Catalog.objects.filter(company=get_request().user.company)
    children = children.difference(Catalog.objects.filter(c_table=Value('{}')))
    return Response(status=status.HTTP_200_OK, data=children.values('id', 'name', 'c_table'))


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated & CatalogPermissions])
def download_file_export_catalog(request, *args, **kwargs):
    url = request.GET.get('url', None)
    prefix_to_remove = URL_CLOUD
    result = url.replace(prefix_to_remove, "")
    data = FileBuilder365.objects.get(file=result)
    return Response(status=status.HTTP_200_OK, data={"url": url})


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated & CatalogPermissions])
def get_file_export_catalog(request, task_id, *args, **kwargs):
    data_attachment = FileBuilder365.objects.filter(task_id=task_id).first()
    url = data_attachment.file.url
    return Response(status=status.HTTP_200_OK, data={"url": url})


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated & CatalogPermissions])
def get_status_process(request, task_id,  *args, **kwargs):
    result = AsyncResult(task_id)

    return Response(status=status.HTTP_200_OK, data={"status": result.status})


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated & CatalogPermissions])
def export_catalog_ver2(request, *args, **kwargs):
    company_id = request.user.company.id
    user_id = request.user.id
    pk = request.GET.get('pk_catalog', None)
    process_export = process_export_catalog.delay(pk, company_id, user_id)
    task_id = process_export.id

    return Response(status=status.HTTP_200_OK, data={"task_id": task_id})


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated & CatalogPermissions])
def export_catalog(request, *args, **kwargs):
    workbook = Workbook()
    pk = request.GET.get('pk_catalog', None)

    if pk is None:
        data_catalog = Catalog.objects.filter(is_ancestor=True, parents=None, company=request.user.company)
        for catalog in data_catalog:
            child_catalogs = Catalog.objects.filter(parents=catalog.id)
            for data_catalog in child_catalogs:
                handle_export(data_catalog.id, workbook, catalog.name)

    else:
        check_catalog = Catalog.objects.get(id=pk)

        if check_catalog.is_ancestor:
            child_catalogs = Catalog.objects.filter(parents=pk)
            for data_catalog in child_catalogs:
                handle_export(data_catalog.id, workbook, check_catalog.name)

        else:
            data_parent_catalog = check_catalog.parents.first()
            handle_export(pk, workbook, data_parent_catalog.name)

    return file_response(workbook=workbook, title='catalog')


def handle_export(pk, workbook, sheet_name):
    root_catalogs = Catalog.objects.filter(id=pk)
    all_paths = []
    name = root_catalogs.first().name
    level = Catalog.objects.get(id=pk).get_ordered_levels()
    # get all path catalog to cost table
    for root_catalog in root_catalogs:
        find_all_paths(root_catalog, [], all_paths)

    # get all data point with catalog on path
    rs = []
    for path in all_paths:
        temp = list()
        temp_catalog_not_dtp = list()
        for idx, data in enumerate(path):
            if isinstance(data, list):
                temp.append(data)
                temp_catalog_not_dtp.append(data)
            elif isinstance(data, Catalog):
                tmp = list()
                if data.data_points.exists():
                    temp_list = []
                    for data_point in data.data_points.all():
                        temp_list.append(data.name)
                        temp_list.append(data.icon)
                        if data_point.unit:
                            temp_list.append(data_point.unit.name)
                        else:
                            temp_list.append('')
                        temp_list.append(data_point.value)
                        temp_list.append(data_point.linked_description)
                        tmp.append(temp_list)
                        temp_list = []
                        if idx != 0 and len(path) == 2:
                            temp_catalog_not_dtp.append([data.id])
                    temp.append(tmp)
                else:
                    if idx != 0:
                        temp.append([data.id])
                        temp_catalog_not_dtp.append([data.id])
        if temp == []:
            rs.append(temp_catalog_not_dtp)

        elif len(temp) == 1 and isinstance(temp[0], list):
            rs.append(temp_catalog_not_dtp)
        else:
            rs.append(temp)

    # get all path with data point
    result_paths = []
    for i in rs:
        generate_paths(i, [], result_paths)

    # handle write headers all catalog and cost table
    # workbook = Workbook()
    data_sheet_name = sheet_name + '-' + name
    data_sheet_name = data_sheet_name.replace("/", "-")
    data_sheet_name = data_sheet_name[:31]
    catalog_sheet = workbook.create_sheet(title=data_sheet_name)
    headers = []
    for i in range(1, len(level) + 1):
        data_level = level[i - 1]
        iteration_headers = [f"{data_level.name}", "image", f"value", f"unit", f"des"]
        headers.extend(iteration_headers)

    all_key = []
    for path in result_paths:
        for data in path:
            if isinstance(data, dict):
                for key in data.keys():
                    if key not in all_key:
                        all_key.append(key)
                break
    headers.extend(all_key)
    catalog_sheet.append(headers)

    # write all data into excel
    if len(level) > 0:
        for path in result_paths:
            number = len(level) * 5
            row = [""] * (number + len(all_key))
            for idx, data in enumerate(path):
                # write catalog data
                if isinstance(data, int):
                    data_catalog = Catalog.objects.get(id=data)
                    if idx == 0:
                        row[idx] = data_catalog.name
                        row[idx + 1] = data_catalog.icon
                    else:
                        row[idx * 5] = data_catalog.name
                        row[idx * 5 + 1] = data_catalog.icon

                elif isinstance(data, list):
                    if idx == 0:
                        row[idx] = data[0]
                        row[idx + 1] = data[1]
                        row[idx + 2] = data[2]
                        row[idx + 3] = data[3]
                        row[idx + 4] = data[4]
                    else:
                        row[idx * 5] = data[0]
                        row[idx * 5 + 1] = data[1]
                        row[idx * 5 + 2] = data[2]
                        row[idx * 5 + 3] = data[3]
                        row[idx * 5 + 4] = data[4]
                # write cost table
                elif isinstance(data, dict):
                    temp_len = len(headers) - number
                    temp_length = headers[-temp_len:]
                    for index, header in enumerate(temp_length, number):
                        if header in data:
                            row[index] = data[header]
                        else:
                            row[index] = ""
                # write catalog not data point
                else:
                    if idx == 0:
                        row[idx] = data
                    else:
                        row[idx * 5] = data
            catalog_sheet.append(row)


def generate_paths(categories, current_path, result):
    if categories == [[]]:
        current_category = []
    elif categories == []:
        current_category = categories
    elif isinstance(categories[0][0], dict):
        current_category = categories[0]

    elif isinstance(categories[0][0], list):

        current_category = categories[0]

    elif isinstance(categories[0][0], int):
        current_category = categories[0]

    else:
        current_category = set(categories[0])
    for item in current_category:
        new_path = current_path + [item]

        if len(categories) == 1:
            result.append(new_path)
        else:
            generate_paths(categories[1:], new_path, result)


def find_all_paths(node, current_path, all_paths):
    if node is None:
        return

    current_path.append(node)
    if not node.children.exists():
        path = list(current_path)
        if hasattr(node, 'c_table') and node.c_table:
            result_list = []
            headers = node.c_table['header']
            for row in node.c_table["data"]:
                result_dict = dict(zip(headers, row))
                result_list.append(result_dict)
            all_paths.append(path + [result_list])

        elif hasattr(node, 'c_table') and not node.c_table:
            # all_paths.append(path + [node.name])
            all_paths.append(path)
        current_path.pop()
    else:
        for child in node.children.all():
            find_all_paths(child, list(current_path), all_paths)

    if len(current_path) > 1:
        current_path.pop()


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated & CatalogPermissions])
def import_catalog_v2(request):
    file = request.FILES['file']
    serializer = FileBuilder365ResSerializer(data={"file": file}, context={'request': request})
    serializer.is_valid(raise_exception=True)
    upload_file = serializer.save()

    company = request.user.company
    process_export = import_catalog_task.delay(upload_file.pk, company.pk, request.user.pk)
    task_id = process_export.id

    return Response(status=status.HTTP_200_OK, data={"task_id": task_id})


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated & CatalogPermissions])
def import_catalog(request):
    file = request.FILES['file']
    serializer = FileBuilder365ResSerializer(data={"file": file}, context={'request': request})
    serializer.is_valid(raise_exception=True)
    upload_file = serializer.save()

    company = request.user.company
    import_catalog_task(upload_file.pk, company.pk, request.user.pk)

    return Response(status=status.HTTP_200_OK)


@api_view(['PUT'])
@permission_classes([permissions.IsAuthenticated & CatalogPermissions])
def delete(request):
    """
    Payload: {"deleted_items": [id], "tree": {"parent_id": [child_id]}, "is_delete_children": boolean}
    """
    data = request.data
    serializer = catalog.DeleteCatalogSerializer(data=data)
    serializer.is_valid(raise_exception=True)
    serializer.save()
    return Response(status=status.HTTP_200_OK)


@api_view(['PUT'])
@permission_classes([permissions.IsAuthenticated & CatalogPermissions])
def move_catalog(request):
    """
    Payload: [id: int]
    """
    data = request.data
    catalog_update = []
    for index, data_catalog in enumerate(data):
        catalog = Catalog.objects.get(pk=data_catalog)
        catalog.index = index
        catalog_update.append(catalog)

    Catalog.objects.bulk_update(catalog_update, ['index'])
    rs = Catalog.objects.filter(id__in=data)
    catalogs = CatalogSerializer(
        rs, many=True, context={'request': request}).data
    return Response(status=status.HTTP_200_OK, data=catalogs)