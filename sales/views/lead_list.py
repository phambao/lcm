import uuid

from django.core.files.base import ContentFile
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django_filters import rest_framework as filters
from django.apps import apps
from openpyxl import Workbook, load_workbook
from rest_framework import generics, permissions
from rest_framework import status, filters as rf_filters
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response

from api.middleware import get_request
from base.permissions import LeadPermissions
from base.utils import file_response
from base.views.base import CompanyFilterMixin
from sales.filters.proposal import PriceComparisonFilter, ProposalWritingFilter
from sales.serializers.proposal import PriceComparisonCompactSerializer, ProposalWritingByLeadSerializer
from ..filters.lead_list import ContactsFilter, ActivitiesFilter, LeadDetailFilter
from ..models.lead_list import LeadDetail, Activities, Contact, PhoneOfContact, Photos, ContactTypeName, \
    ProjectType, TagLead, PhaseActivity, TagActivity, SourceLead
from ..serializers import lead_list
from ..serializers.lead_list import PhotoSerializer, LeadDetailCreateSerializer

PASS_FIELDS = ['user_create', 'user_update', 'lead']
LEAD_FIELDS = ('activities', 'contacts', 'contacts__phone_contacts', 'project_types', 'salesperson',
               'sources', 'tags', 'photos')


class LeadDetailList(CompanyFilterMixin, generics.ListCreateAPIView):
    queryset = LeadDetail.objects.all().prefetch_related(*LEAD_FIELDS)
    serializer_class = lead_list.LeadDetailCreateSerializer
    permission_classes = [permissions.IsAuthenticated & LeadPermissions]
    filter_backends = (filters.DjangoFilterBackend, rf_filters.SearchFilter)
    filterset_class = LeadDetailFilter
    search_fields = ['lead_title', 'street_address', 'notes']


class LeadWithChangeOrderList(CompanyFilterMixin, generics.ListAPIView):
    queryset = LeadDetail.objects.filter(proposals__change_orders__isnull=False).prefetch_related(*LEAD_FIELDS).distinct()
    serializer_class = lead_list.LeadFilterChangeOrderSerializer
    permission_classes = [permissions.IsAuthenticated & LeadPermissions]
    filter_backends = (filters.DjangoFilterBackend, rf_filters.SearchFilter)
    filterset_class = LeadDetailFilter
    search_fields = ['lead_title', 'street_address', 'notes']


class LeadWithProposal(LeadWithChangeOrderList):
    queryset = LeadDetail.objects.filter(proposals__isnull=False).prefetch_related(*LEAD_FIELDS).distinct()


class LeadWithInvoice(LeadWithChangeOrderList):
    queryset = LeadDetail.objects.filter(proposals__invoices__isnull=False).prefetch_related(*LEAD_FIELDS).distinct()


class LeadWithInvoicePayment(LeadWithChangeOrderList):
    queryset = LeadDetail.objects.filter(
        proposals__invoices__payment_histories__isnull=False
    ).prefetch_related(*LEAD_FIELDS).distinct()


class LeadEventList(CompanyFilterMixin, generics.ListAPIView):
    queryset = LeadDetail.objects.all().prefetch_related('schedule_event_lead_list')
    serializer_class = lead_list.LeadViewEventSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = (filters.DjangoFilterBackend, rf_filters.SearchFilter)
    filterset_class = LeadDetailFilter
    search_fields = ['lead_title', 'street_address', 'notes']


class LeadDetailGeneric(generics.RetrieveUpdateDestroyAPIView):
    queryset = LeadDetail.objects.all()
    serializer_class = lead_list.LeadDetailCreateSerializer
    permission_classes = [permissions.IsAuthenticated & LeadPermissions]

    def get_serializer_context(self):
        data = super().get_serializer_context()
        data['pk_lead'] = self.kwargs.get('pk')
        return data

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        if not instance.number_of_click:
            instance.number_of_click = 0
        instance.number_of_click += 1
        instance.recent_click = timezone.now()
        instance.save()
        return Response(serializer.data)


class LeadActivitiesViewSet(generics.ListCreateAPIView):
    serializer_class = lead_list.ActivitiesSerializer
    permission_classes = [permissions.IsAuthenticated & LeadPermissions]
    filter_backends = (filters.DjangoFilterBackend, rf_filters.SearchFilter)
    filterset_class = ActivitiesFilter
    search_fields = ['title', 'phase', 'tag', 'status', 'assigned_to']

    def get_queryset(self):
        get_object_or_404(LeadDetail.objects.filter(company=get_request().user.company), pk=self.kwargs['pk_lead'])
        return Activities.objects.filter(lead_id=self.kwargs['pk_lead'])


class LeadActivitiesDetailViewSet(generics.RetrieveUpdateDestroyAPIView):
    queryset = Activities.objects.all()
    serializer_class = lead_list.ActivitiesSerializer
    permission_classes = [permissions.IsAuthenticated & LeadPermissions]

    def put(self, request, *args, **kwargs):
        data = request.data
        [data.pop(field) for field in PASS_FIELDS if field in data]
        instance = self.get_object()
        instance.user_update = request.user
        instance.save()
        return super().put(request, *args, **kwargs)


class LeadPhotosViewSet(generics.ListCreateAPIView):
    serializer_class = lead_list.PhotoSerializer
    permission_classes = [permissions.IsAuthenticated & LeadPermissions]

    def get_queryset(self):
        get_object_or_404(LeadDetail.objects.filter(company=get_request().user.company), pk=self.kwargs['pk_lead'])
        return Photos.objects.filter(lead_id=self.kwargs['pk_lead'])


class LeadPhotosDetailViewSet(generics.RetrieveDestroyAPIView):
    """
    Used for get params
    """
    queryset = Photos.objects.all()
    serializer_class = lead_list.PhotoSerializer
    permission_classes = [permissions.IsAuthenticated & LeadPermissions]

    def delete(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class ContactsViewSet(CompanyFilterMixin, generics.ListCreateAPIView):
    queryset = Contact.objects.all()
    serializer_class = lead_list.ContactsSerializer
    permission_classes = [permissions.IsAuthenticated & LeadPermissions]
    filter_backends = (filters.DjangoFilterBackend,)
    filterset_class = ContactsFilter


class LeadNoContactsViewSet(generics.ListAPIView):
    queryset = Contact.objects.all()
    serializer_class = lead_list.ContactsSerializer
    permission_classes = [permissions.IsAuthenticated & LeadPermissions]
    filter_backends = (filters.DjangoFilterBackend, rf_filters.SearchFilter)
    filterset_class = ContactsFilter
    search_fields = ['first_name', 'last_name']

    def get_queryset(self):
        return Contact.objects.filter(company=get_request().user.company).exclude(
            leads=self.kwargs['pk_lead']).distinct()


class ContactsDetailViewSet(generics.RetrieveUpdateDestroyAPIView):
    queryset = Contact.objects.all()
    serializer_class = lead_list.ContactsSerializer
    permission_classes = [permissions.IsAuthenticated & LeadPermissions]


class PhoneOfContactsViewSet(generics.ListCreateAPIView):
    serializer_class = lead_list.PhoneContactsSerializer
    permission_classes = [permissions.IsAuthenticated & LeadPermissions]

    def get_queryset(self):
        get_object_or_404(Contact.objects.filter(company=get_request().user.company), pk=self.kwargs['pk_contact'])
        return PhoneOfContact.objects.filter(contact_id=self.kwargs['pk_contact'])


class LeadContactsViewSet(generics.ListCreateAPIView):
    queryset = Contact.objects.all()
    serializer_class = lead_list.ContactsSerializer
    permission_classes = [permissions.IsAuthenticated & LeadPermissions]
    filter_backends = (filters.DjangoFilterBackend, rf_filters.SearchFilter)
    filterset_class = ContactsFilter
    search_fields = ['first_name', 'last_name', 'email', 'phone_contacts__phone_number']

    def get_queryset(self):
        get_object_or_404(LeadDetail.objects.filter(company=get_request().user.company), pk=self.kwargs['pk_lead'])
        return Contact.objects.filter(leads__id=self.kwargs['pk_lead'])


class LeadContactDetailsViewSet(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = lead_list.ContactsSerializer
    permission_classes = [permissions.IsAuthenticated & LeadPermissions]

    def get_queryset(self):
        try:
            qs = LeadDetail.objects.get(pk=self.kwargs['pk_lead']).contacts.all()
            return qs
        except KeyError:
            pass
        return Contact.objects.all()


class PhoneOfContactsDetailViewSet(generics.RetrieveUpdateDestroyAPIView):
    queryset = PhoneOfContact.objects.all()
    serializer_class = lead_list.PhoneContactsSerializer
    permission_classes = [permissions.IsAuthenticated & LeadPermissions]


class ContactTypeNameGenericView(CompanyFilterMixin, generics.ListCreateAPIView):
    queryset = ContactTypeName.objects.all()
    serializer_class = lead_list.ContactTypeNameSerializer
    permission_classes = [permissions.IsAuthenticated & LeadPermissions]


class ContactTypeNameDetailGenericView(generics.RetrieveUpdateDestroyAPIView):
    queryset = ContactTypeName.objects.all()
    serializer_class = lead_list.ContactTypeNameSerializer
    permission_classes = [permissions.IsAuthenticated & LeadPermissions]


class ProjectTypeGenericView(CompanyFilterMixin, generics.ListCreateAPIView):
    queryset = ProjectType.objects.all()
    serializer_class = lead_list.ProjectTypeSerializer
    permission_classes = [permissions.IsAuthenticated & LeadPermissions]


class ProjectTypeDetailGenericView(generics.RetrieveUpdateDestroyAPIView):
    queryset = ProjectType.objects.all()
    serializer_class = lead_list.ProjectTypeSerializer
    permission_classes = [permissions.IsAuthenticated & LeadPermissions]


class TagLeadGenericView(CompanyFilterMixin, generics.ListCreateAPIView):
    queryset = TagLead.objects.all()
    serializer_class = lead_list.TagLeadSerializer
    permission_classes = [permissions.IsAuthenticated & LeadPermissions]


class TagLeadDetailGenericView(generics.RetrieveUpdateDestroyAPIView):
    queryset = TagLead.objects.all()
    serializer_class = lead_list.TagLeadSerializer
    permission_classes = [permissions.IsAuthenticated & LeadPermissions]


class TagActivitiesGenericView(generics.ListCreateAPIView):
    queryset = TagActivity.objects.all()
    serializer_class = lead_list.TagActivitySerializer
    permission_classes = [permissions.IsAuthenticated & LeadPermissions]


class TagActivityDetailGenericView(generics.RetrieveUpdateDestroyAPIView):
    queryset = TagActivity.objects.all()
    serializer_class = lead_list.TagActivitySerializer
    permission_classes = [permissions.IsAuthenticated & LeadPermissions]


class PhaseActivitiesGenericView(generics.ListCreateAPIView):
    queryset = PhaseActivity.objects.all()
    serializer_class = lead_list.PhaseActivitySerializer
    permission_classes = [permissions.IsAuthenticated & LeadPermissions]


class PhaseActivityDetailGenericView(generics.RetrieveUpdateDestroyAPIView):
    queryset = PhaseActivity.objects.all()
    serializer_class = lead_list.PhaseActivitySerializer
    permission_classes = [permissions.IsAuthenticated & LeadPermissions]


class SourceLeadGenericView(CompanyFilterMixin, generics.ListCreateAPIView):
    queryset = SourceLead.objects.all()
    serializer_class = lead_list.SourceLeadSerializer
    permission_classes = [permissions.IsAuthenticated & LeadPermissions]


class SourceLeadDetailGenericView(generics.RetrieveUpdateDestroyAPIView):
    queryset = SourceLead.objects.all()
    serializer_class = lead_list.SourceLeadSerializer
    permission_classes = [permissions.IsAuthenticated & LeadPermissions]


class PriceComparisonByLeadViewSet(generics.ListAPIView):
    serializer_class = PriceComparisonCompactSerializer
    permission_classes = [permissions.IsAuthenticated & LeadPermissions]
    filter_backends = (filters.DjangoFilterBackend, rf_filters.SearchFilter)
    filterset_class = PriceComparisonFilter
    search_fields = ('name',)

    def get_queryset(self):
        model = apps.get_model('sales', 'PriceComparison')
        try:
            qs = model.objects.filter(lead=self.kwargs['pk_lead'], company=get_request().user.company)
            return qs
        except KeyError:
            pass
        return model.objects.none()


class ProposalWritingByLeadViewSet(generics.ListAPIView):
    serializer_class = ProposalWritingByLeadSerializer
    permission_classes = [permissions.IsAuthenticated & LeadPermissions]
    filter_backends = (filters.DjangoFilterBackend, rf_filters.SearchFilter)
    filterset_class = ProposalWritingFilter
    search_fields = ('name',)

    def get_queryset(self):
        model = apps.get_model('sales', 'ProposalWriting')
        try:
            qs = model.objects.filter(lead=self.kwargs['pk_lead'], company=get_request().user.company)
            return qs
        except KeyError:
            pass
        return model.objects.none()


@api_view(['DELETE'])
@permission_classes([permissions.IsAuthenticated & LeadPermissions])
def delete_activities(request, pk_lead):
    """
        DELETE: delete multiple activities
    """

    if request.method == 'DELETE':
        ids = request.data
        activities = Activities.objects.filter(id__in=ids, lead=pk_lead)
        activities.delete()
    return Response(status=status.HTTP_204_NO_CONTENT)


@api_view(['DELETE'])
@permission_classes([permissions.IsAuthenticated & LeadPermissions])
def delete_leads(request):
    """
        DELETE: delete multiple leads
    """

    if request.method == 'DELETE':
        ids = request.data
        leads = LeadDetail.objects.filter(id__in=ids)
        leads.delete()
    return Response(status=status.HTTP_204_NO_CONTENT)


@api_view(['DELETE'])
@permission_classes([permissions.IsAuthenticated & LeadPermissions])
def delete_contacts(request):
    """
        DELETE: delete multiple contacts
    """

    if request.method == 'DELETE':
        ids = request.data
        contacts = Contact.objects.filter(id__in=ids)
        contacts.delete()
    return Response(status=status.HTTP_204_NO_CONTENT)


@api_view(['PUT'])
@permission_classes([permissions.IsAuthenticated & LeadPermissions])
def unlink_contact_from_lead(request, pk_lead):
    """
        PUT: unlink contact from lead
    """

    if request.method == 'PUT':
        contact_ids = request.data
        lead = LeadDetail.objects.get(pk=pk_lead)
        contacts_to_unlink = lead.contacts.all().filter(id__in=contact_ids)
        lead.contacts.remove(*contacts_to_unlink)
        data = lead_list.ContactsSerializer(
            contacts_to_unlink, many=True, context={'request': request}).data
    return Response(status=status.HTTP_200_OK, data=data)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def get_summaries(request):
    data = {
        'closed_job': {'title': 'Total $ of Proposals Sent & Awaiting Approval',
                       'number': '$5,000,000',
                       'content': 200},
        'number_of_job': {'title': 'Proposal Win Ratio',
                          'number': '80%',
                          'content': 200},
        'closed_ratio': {'title': '$ of Projects Awarded in Last 90 Days',
                         'number': '$5,500,000',
                         'content': -200},
        'dollar_of_job': {'title': 'Total Awarded Projects YTD',
                          'number': '$15,000,000',
                          'content': 2000},
    }
    if request.method == 'GET':
        pass
    return Response(status=status.HTTP_200_OK, data=data)


@api_view(['PUT'])
@permission_classes([permissions.IsAuthenticated & LeadPermissions])
def link_contacts_to_lead(request, pk_lead):
    """
        PUT: link contacts to lead
    """

    if request.method == 'PUT':
        contact_ids = request.data
        lead = LeadDetail.objects.get(pk=pk_lead)
        contacts_to_link = Contact.objects.filter(id__in=contact_ids)
        lead.contacts.add(*contacts_to_link)
        data = lead_list.ContactsSerializer(
            contacts_to_link, many=True, context={'request': request}).data
    return Response(status=status.HTTP_200_OK, data=data)


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated & LeadPermissions])
def upload_multiple_photo(request, pk_lead):
    try:
        files = request.FILES.getlist('files')
    except KeyError:
        return Response(status=status.HTTP_400_BAD_REQUEST, data={"message": "File not found"})
    photo_id = []
    for file in files:
        file_name = uuid.uuid4().hex + '.' + file.name.split('.')[-1]
        content_file = ContentFile(file.read(), name=file_name)

        # is needed to bulk_create?
        photo = Photos.objects.create(photo=content_file, user_create=request.user,
                                      user_update=request.user, lead_id=pk_lead)
        photo_id.append(photo.id)
    photos = Photos.objects.filter(pk__in=photo_id)
    serializer = PhotoSerializer(photos, many=True)

    return Response(serializer.data)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated & LeadPermissions])
def export_data(request):
    workbook = Workbook()

    # Mô hình LeadDetail
    lead_detail_sheet = workbook.create_sheet(title='LeadDetail')
    lead_details = LeadDetail.objects.all()
    lead_detail_fields = ['Lead Title', 'Street Address', 'Country', 'City', 'State', 'Zip Code',
                          'Status', 'Proposal Status', 'Notes', 'Confidence', 'Estimate Revenue From',
                          'Estimate Revenue To', 'Project Types', 'Sales Person', 'Sources', 'Tags', 'Number of Click',
                          'Projected Sale Date']
    lead_detail_sheet.append(lead_detail_fields)
    for index, lead_detail in enumerate(lead_details, 1):
        projected_sale_date = lead_detail.projected_sale_date
        if projected_sale_date is not None:
            projected_sale_date = projected_sale_date.replace(tzinfo=None)
        else:
            projected_sale_date = ''
        row_data = [
            lead_detail.lead_title, lead_detail.street_address,
            lead_detail.country,
            lead_detail.city,
            lead_detail.state,
            lead_detail.zip_code, lead_detail.status,
            lead_detail.proposal_status, lead_detail.notes,
            lead_detail.confidence, lead_detail.estimate_revenue_from,
            lead_detail.estimate_revenue_to,
            ', '.join(str(pt.id) for pt in lead_detail.project_types.all()),
            ', '.join(str(salesperson.username) for salesperson in lead_detail.salesperson.all()),
            ', '.join(str(source.id) for source in lead_detail.sources.all()),
            ', '.join(str(tag.id) for tag in lead_detail.tags.all()),
            lead_detail.number_of_click, projected_sale_date
        ]

        lead_detail_sheet.append(row_data)

    # Mô hình ProjectType
    project_type_sheet = workbook.create_sheet(title='ProjectType')
    project_types = ProjectType.objects.all()
    project_type_fields = ['Project Type Id', 'Project Type Name', 'User Create', 'User Update', 'Created Date']
    project_type_sheet.append(project_type_fields)
    for project_type in project_types:
        create_date = project_type.created_date
        if create_date is not None:
            create_date = create_date.replace(tzinfo=None)
        else:
            create_date = ''

        row = [
            project_type.id, project_type.name,
            project_type.user_create.id if project_type.user_create else '',
            project_type.user_update.id if project_type.user_update else '',
            create_date
        ]
        project_type_sheet.append(row)

    # Mô hình Source
    source_lead_sheet = workbook.create_sheet(title='SourceLead')
    source_leads = SourceLead.objects.all()
    source_lead_fields = ['Source Lead Id', 'Source Lead Name', 'User Create', 'User Update', 'Created Date']
    source_lead_sheet.append(source_lead_fields)
    for source_lead in source_leads:
        create_date = source_lead.created_date
        if create_date is not None:
            create_date = create_date.replace(tzinfo=None)
        else:
            create_date = ''

        row = [
            source_lead.id, source_lead.name,
            source_lead.user_create.id if source_lead.user_create else '',
            source_lead.user_update.id if source_lead.user_update else '',
            create_date
        ]
        source_lead_sheet.append(row)

    # Mô hình contacts
    contacts_sheet = workbook.create_sheet(title='Contact')
    contacts = Contact.objects.all()
    contact_fields = ['Contact Id', 'First Name', 'Last Name', 'Gender', 'Email', 'Street',
                      'City', 'State', 'Country', 'Zip Code', 'Lead', 'User Create', 'User Update',
                      'Created Date']
    contacts_sheet.append(contact_fields)
    for contact in contacts:
        create_date = contact.created_date
        if create_date is not None:
            create_date = create_date.replace(tzinfo=None)
        else:
            create_date = ''

        row = [
            contact.id, contact.first_name, contact.last_name, contact.gender, contact.email,contact.street,
            contact.city,
            contact.state,
            contact.country,
            contact.zip_code,
            ', '.join(str(ct.id)for ct in contact.leads.all()),
            contact.user_create.id if contact.user_create else '',
            contact.user_update.id if contact.user_update else '',
            create_date
        ]
        contacts_sheet.append(row)

    # Mô hình Tag
    tag_lead_sheet = workbook.create_sheet(title='TagLead')
    tag_leads = TagLead.objects.all()
    tag_lead_fields = ['Tag Lead Id', 'Tag Lead Name', 'User Create', 'User Update', 'Created Date']
    tag_lead_sheet.append(tag_lead_fields)
    for tag_lead in tag_leads:
        create_date = tag_lead.created_date
        if create_date is not None:
            create_date = create_date.replace(tzinfo=None)
        else:
            create_date = ''

        row = [
            tag_lead.id, tag_lead.name,
            tag_lead.user_create.id if tag_lead.user_create else '',
            tag_lead.user_update.id if tag_lead.user_update else '',
            create_date
        ]
        tag_lead_sheet.append(row)

    # Mô hình Photo
    photo_sheet = workbook.create_sheet(title='Photos')
    photos = Photos.objects.all()
    photo_fields = ['Photo Id', 'Photo', 'lead_title']
    photo_sheet.append(photo_fields)
    photo_serializer = PhotoSerializer(photos, many=True)
    for photo in photos:
        row = [
            photo.id, photo.photo.name, photo.lead.lead_title
        ]
        photo_sheet.append(row)

    return file_response(workbook=workbook, title='Lead_detail')


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated & LeadPermissions])
def import_data(request):
    file = request.FILES['file']

    workbook = load_workbook(file)
    lead_detail_sheet = workbook['LeadDetail']
    project_type_sheet = workbook['ProjectType']
    source_lead_sheet = workbook['SourceLead']
    tag_lead_sheet = workbook['TagLead']
    contacts_sheet = workbook['Contact']
    photo_sheet = workbook['Photos']
    # Import dữ liệu cho mô hình LeadDetail
    lead_detail_data = []
    max_row = lead_detail_sheet.max_row
    temp_rs = []
    # for row in lead_detail_sheet.iter_rows(min_row=2, values_only=True):
    for row_number in range(max_row, 1, -1):
        row = lead_detail_sheet[row_number]
        notes = row[8].value
        street_address = row[1].value
        zip_code = row[5].value
        if notes is None:
            notes = ''
        if street_address is None:
            street_address = ''
        if zip_code is None:
            zip_code = ''
        data_create = {
            'lead_title': row[0].value,
            'street_address': street_address,
            'country': row[2].value,
            'city': row[3].value,
            'state': row[4].value,
            'zip_code': zip_code,
            'status': row[6].value,
            'proposal_status': row[7].value,
            'notes': notes,
            'confidence': row[9].value,
            'estimate_revenue_from': row[10].value,
            'estimate_revenue_to': row[11].value,
            'number_of_click': row[16].value,
            'projected_sale_date': row[17].value
        }
        ld = LeadDetail.objects.create(**data_create)
        temp_rs.append(ld)
        data_tag = row[15].value
        if data_tag:
            data_tag = data_tag.split(',')
            tags = []
            for tag in data_tag:
                tag = tag.strip()
                try:
                    temp = TagLead.objects.get(name=tag, company=request.user.company)
                    tags.append(temp)
                except TagLead.DoesNotExist:
                    data_create = TagLead.objects.create(**{"name": tag})
                    tags.append(data_create)

            ld.tags.add(*tags)

        pts = []
        data_project_type = row[12].value
        if data_project_type:
            data_project_type = data_project_type.split(',')
            for project_type in data_project_type:
                project_type = project_type.strip()
                try:
                    temp = ProjectType.objects.get(name=project_type, company=request.user.company)
                    pts.append(temp)
                except ProjectType.DoesNotExist:
                    data_create = ProjectType.objects.create(**{"name": project_type})
                    pts.append(data_create)

            ld.project_types.add(*pts)

        sources = []
        data_sources = row[14].value
        if data_sources:
            data_sources = data_sources.split(',')
            for source in data_sources:
                source = source.strip()
                try:
                    temp = SourceLead.objects.get(name=source, company=request.user.company)
                    sources.append(temp)
                except SourceLead.DoesNotExist:
                    data_create = SourceLead.objects.create(**{"name": source})
                    sources.append(data_create)

            ld.sources.add(*sources)

    # max_row = photo_sheet.max_row
    # url = "https://cdn-lcm-staging.sfo3.cdn.digitaloceanspaces.com/lcm-staging/2023/09/08/83e1d22ee06d4f2f9f9891c0d029a91e.jpg"
    # # for row in lead_detail_sheet.iter_rows(min_row=2, values_only=True):
    # attachment_create = []
    # for row_number in range(max_row, 1, -1):
    #     row = photo_sheet[row_number]
    #     try:
    #         # Gửi yêu cầu GET để tải file từ URL
    #         response = requests.get(url)
    #         response.raise_for_status()  # Kiểm tra trạng thái của yêu cầu
    #
    #         # Lấy nội dung của file từ phản hồi
    #         file_content = response.content
    #         content_type = response.headers.get("content-type")
    #         file_in_memory = io.BytesIO(file_content)
    #         file_extension = mimetypes.guess_extension(content_type)
    #         file_in_memory.name = uuid.uuid4().hex + file_extension
    #         a = file_in_memory.name
    #         size = response.headers["Content-Length"]
    #         content_file = ContentFile(file.read(), name=file_in_memory.name)
    #         attachment = FileBuilder365(
    #             file=content_file,
    #             user_create=request.user,
    #             user_update=request.user,
    #             name=file.name,
    #             size=size
    #         )
    #         attachment_create.append(attachment)
    #         # Tiếp theo, bạn có thể làm việc với file_content theo ý muốn.
    #         # Ví dụ: Lưu nó vào một biến hoặc thực hiện các xử lý khác.
    #
    #         # Ví dụ: Lưu nó vào một biến
    #         # file_in_memory = file_content
    #
    #         # Sau khi đã lưu file vào bộ nhớ RAM, bạn có thể thao tác với nó theo ý muốn.
    #         # Ví dụ: In độ dài của file_content
    #
    #     except requests.exceptions.RequestException as e:
    #         print(f"Lỗi khi tải file: {str(e)}")
    # FileBuilder365.objects.bulk_create(attachment_create)
    rs = LeadDetailCreateSerializer(
        temp_rs, many=True, context={'request': request}).data
    return Response(status=status.HTTP_200_OK, data=rs)
