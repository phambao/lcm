import zipfile
from datetime import datetime
from io import BytesIO

from celery import shared_task, current_task
from django.contrib.auth import get_user_model
from django.contrib.contenttypes.models import ContentType
from django.core.mail import send_mail, EmailMessage
from django.http import HttpRequest
from django.core.files.base import ContentFile
from django.apps import apps
from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook

from api.middleware import set_request
from api.models import ActivityLog, CompanyBuilder
from base.models.config import FileBuilder365
from base.utils import str_to_class
from base.constants import null, true, false
from sales.models import Catalog, UnitLibrary, CatalogLevel, DataPoint


@shared_task()
def process_export_catalog(pk, company, user_id):
    from sales.views.catalog import handle_export
    workbook = Workbook(write_only=True)
    task_id = current_task.request.id
    if pk is None:
        data_catalog = Catalog.objects.filter(is_ancestor=True, parents=None, company=company)
        for catalog in data_catalog:
            child_catalogs = Catalog.objects.filter(parents=catalog.id)
            for data_catalog in child_catalogs:
                handle_export(data_catalog.id, workbook, catalog.name)


    else:
        check_catalog = Catalog.objects.get(id=pk)
        if check_catalog.is_ancestor:
            child_catalogs = Catalog.objects.filter(parents=pk)
            for data_catalog in child_catalogs:
                handle_export(data_catalog.id, workbook, check_catalog.name)

        else:
            data_parent_catalog = check_catalog.parents.first()
            handle_export(pk, workbook, data_parent_catalog.name)

    bytes_io = BytesIO()
    workbook.save(bytes_io)
    bytes_io.seek(0)
    current_datetime = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"catalog_{current_datetime}.xlsx"
    handle_save_file(bytes_io, filename, user_id, task_id)


@shared_task()
def export_proposal(list_proposal, user_id):
    zip_bytes = BytesIO()
    rs = []
    task_id = current_task.request.id
    for proposal in list_proposal:
        wb = Workbook()
        formula_sheet = wb.create_sheet('Formula')
        data_entry_sheet = wb.create_sheet('Data Entry')
        data_view_sheet = wb.create_sheet('Data View')
        ProposalWriting = apps.get_model('sales', 'ProposalWriting')
        proposal = ProposalWriting.objects.get(pk=proposal)
        estimates = proposal.get_estimates()
        formula_sheet.append(['Estimate', 'Formula', 'Material', 'Quantity', 'Unit', 'Unit Cost',
                              'Total Cost', 'Markup', 'Margin', 'Charge'])
        data_view_sheet.append(['Estimate', 'Name', 'Unit', 'Price'])
        data_entry_sheet.append(['Estimate', 'Name', 'Unit', 'Value'])
        for estimate in estimates:
            # formula sheet
            formulas = estimate.get_formula()
            for formula in formulas:
                material = formula.material or "{}"
                formula_sheet.append(
                    [estimate.name, formula.name, eval(material).get('name'), formula.quantity,
                     formula.unit, formula.unit_price, formula.total_cost, formula.markup, formula.margin,
                     formula.charge])

            # Data view
            for data_view in estimate.data_views.all():
                unit = data_view.unit.name if data_view.unit else None
                data_view_sheet.append([estimate.name, data_view.name, unit, data_view.result])

            for data_entry in estimate.data_entries.all():
                data_entry_sheet.append([estimate.name, data_entry.data_entry.name,
                                         data_entry.get_unit(), data_entry.get_value()])

        default_sheet = wb.active
        wb.remove(default_sheet)
        bytes_io = BytesIO()
        wb.save(bytes_io)
        bytes_io.seek(0)
        temp = {'proposal_name': proposal.name, 'work_book': bytes_io}
        rs.append(temp)

    if len(rs) > 1:
        with zipfile.ZipFile(zip_bytes, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for data in rs:
                name = data['proposal_name']
                zipf.writestr(f'{name}.xlsx', data['work_book'].getvalue())

        zip_bytes.seek(0)
        current_datetime = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"proposal_{current_datetime}.zip"
        handle_save_file(zip_bytes, filename, user_id, task_id)

    else:
        proposal_name = rs[0]['proposal_name']
        file_name = f"{proposal_name}.xlsx"
        handle_save_file(rs[0]['work_book'], file_name, user_id, task_id)


def handle_save_file(bytes_io, file_name, user_id, task_id):
    content = ContentFile(bytes_io.read())
    content.seek(0)

    attachment = FileBuilder365()
    user = get_user_model().objects.get(pk=user_id)
    request = HttpRequest()
    request.user = user
    set_request(request)

    attachment.file.save(file_name, content)
    attachment.size = content.size
    attachment.name = file_name
    attachment.task_id = task_id
    attachment.save()


@shared_task()
def celery_send_mail(subject, message, from_email, recipient_list,
                     fail_silently=False, auth_user=None, auth_password=None,
                     connection=None, html_message=None, file=None):
    send_mail(subject, message, from_email, recipient_list,
              fail_silently=fail_silently, auth_user=auth_user, auth_password=auth_password,
              connection=connection, html_message=html_message)


@shared_task(name='send-email')
def send_mail_with_attachment(subject, body, from_email, to, attachments, bcc=None, connection=None,
                              headers=None, cc=None, reply_to=None):
    email = EmailMessage(subject, body, from_email, to, bcc=bcc, connection=connection,
                         attachments=None, headers=headers, cc=cc, reply_to=reply_to)

    for attachment in attachments:
        email.attach(attachment.name, attachment.read())
    email.send()


@shared_task()
def activity_log(model, instance, action, serializer_name, base_import_file, user_id):
    """
    Parameters:
        model: int
        instance: pk
        action: int
        serializer_name: str
        base_import_file: str
    """
    user = get_user_model().objects.get(pk=user_id)
    request = HttpRequest()
    request.user = user
    set_request(request)

    content_type = ContentType.objects.get_for_id(model)
    model_class = content_type.model_class()
    instance = model_class.objects.get(pk=instance)
    serializer = str_to_class(base_import_file, serializer_name)
    data = serializer(instance).data
    ActivityLog.objects.create(content_type=content_type, content_object=instance, object_id=instance.pk,
                               action=action, last_state=data, next_state={})


@shared_task()
def import_catalog_task(file_pk, company_pk, user_pk):
    company = CompanyBuilder.objects.get(pk=company_pk)
    file = FileBuilder365.objects.get(pk=file_pk)
    request = HttpRequest()
    request.user = get_user_model().objects.get(pk=user_pk)
    set_request(request)
    workbook = load_workbook(file.file, read_only=True)
    for idx, sheetname in enumerate(workbook.sheetnames):
        catalog_sheet = workbook[sheetname]
        parent = None
        ancestor_name, parent_name = sheetname.split('-', 1)
        ancestor = Catalog.objects.get_or_create(name=ancestor_name, company=company, is_ancestor=True)[0]
        try:
            parent = ancestor.children.get(name=parent_name)
        except Catalog.DoesNotExist:
            parent = Catalog.objects.create(name=parent_name, company=company)
            parent.parents.add(ancestor)
        except Catalog.MultipleObjectsReturned:
            parent = ancestor.children.filter(name=parent_name).first()

        for row in catalog_sheet.iter_rows(min_row=0, max_row=1, values_only=True):
            header = row
        length_level, length_cost_table, levels, c_table_header = count_level(header, parent)
        c_table_header = ['name', 'unit', 'cost', *c_table_header[3:]]

        cost_table_cache = set()
        unit = set()
        for row in catalog_sheet.iter_rows(min_row=2, values_only=True):
            unit.add(create_catalog_by_row(row, length_level, company, parent, levels, c_table_header, cost_table_cache))

        unit_library = set(UnitLibrary.objects.filter(company=company).values_list('name', flat=True))
        UnitLibrary.objects.bulk_create([UnitLibrary(name=i, company=company) for i in unit.difference(unit_library) if i])
    file.file.delete(save=True)
    file.delete()


def count_level(header, level_catalog):
    """
    Parameters:
        header: tuple
        level_catalog: catalog object
    """
    length = len(header)
    length_level = 0
    for i, element in enumerate(header):
        if element == 'name':
            length_level = i
    else:
        # in case of no cost table
        if not length_level:
            length_level = i + 1
    parent = None
    levels = []
    level_column_number = int(length_level/5)
    level_query = level_catalog.all_levels.all()

    if level_query.count() == level_column_number:
        levels = level_catalog.get_ordered_levels()
        c_table_header = header[level_column_number*5:]
    elif not level_query.count():
        for i in range(level_column_number):
            parent = CatalogLevel.objects.create(name=header[i*5], parent=parent, catalog=level_catalog)
            levels.append(parent)
        else:
            c_table_header = header[i*5 + 5:]
    else:
        return 0, 0, [], []
    return level_column_number, length - length_level, levels, c_table_header


def create_catalog_by_row(row, length_level, company, root, levels, level_header, cache=set()):
    """
    Parameters:
        row: tuple
        length_level: int
        company: company object
        root: catalog object
        levels: list level objects
    """
    unit = None
    parent = root
    for i in range(length_level):
        name = row[i*5]
        if not name:
            continue
        icon = row[i*5 + 1]

        if parent:
            catalog = parent.children.get_or_create(name=name, company=company, level=levels[i], level_index=i)
            catalog = catalog[0]
            catalog.parents.add(parent)
            if icon:
                catalog.icon = icon
                catalog.save(update_fields=['icon'])
        else:
            catalog = Catalog.objects.create(name=name, company=company)

        unit = None
        if row[i*5 + 3]:
            unit = UnitLibrary.objects.get_or_create(name=row[i*5 + 3], company=company)
            unit = unit[0]

        data_point = {'value': row[i*5 + 2] or '', 'unit': unit, 'linked_description': row[i*5 + 4] or ''}
        DataPoint.objects.get_or_create(catalog=catalog, **data_point)
        parent = catalog
    else:
        if length_level:
            c_table = parent.c_table or {'header': level_header, 'data': []}
            if parent.id not in cache:
                #  Get all data from dictionary except key 'data' and 'header'
                c_table = {key: value for key, value in parent.c_table.items() if key not in ['data', 'header']} if parent.c_table else {}
                c_table['header'] = level_header
                c_table['data'] = []
            # Validate cost table data
            data_create = ['' if value is None else str(value) for value in row[i*5 + 5:]]
            if any(row[i*5 + 5:]):
                c_table['data'].append(data_create)
                parent.c_table = c_table
                parent.save()
            # No cost table
            cache.add(parent.id)
            try:
                unit = row[i*5 + 6]
            except IndexError:
                return
    return unit
